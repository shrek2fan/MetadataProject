import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import argparse
import os

# Define fill styles for highlighting mistakes
highlight_fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
highlight_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Load the approved SUBJECT_LCSH vocabulary with separate English and Spanish terms
def load_approved_subjects(vocabulary_file):
    approved_df = pd.read_excel(vocabulary_file, usecols=[0, 1])  # Load only columns A and B
    approved_subjects = {
        "spanish": set(approved_df.iloc[:, 0].dropna().str.strip()),  # Spanish terms from column A
        "english": set(approved_df.iloc[:, 1].dropna().str.strip()),  # English terms from column B
    }
    print("Debug: Approved subjects loaded:", approved_subjects)  # Log loaded terms for confirmation
    return approved_subjects


# Load the city dataset into a dictionary structure with safe handling for non-string values
def load_city_data(city_dataset_path):
    city_data = pd.read_excel(city_dataset_path)

    def safe_strip(value):
        return str(value).strip() if isinstance(value, str) else ''

    city_info = {}
    for _, row in city_data.iterrows():
        city_info[(safe_strip(row['ES_City']).lower(), 'spanish')] = {
            'country': safe_strip(row['ES_Country']),
            'state': safe_strip(row['ES_State']),
            'coordinates': safe_strip(row["CITIES' LAT_LONG COORDINATES"])
        }
        city_info[(safe_strip(row['EN_City']).lower(), 'english')] = {
            'country': safe_strip(row['EN_Country']),
            'state': safe_strip(row['EN_State']),
            'coordinates': safe_strip(row["CITIES' LAT_LONG COORDINATES"])
        }
    return city_info


# List of valid series names
series_values = [
    'Martin Amador, 1856-1904',
    'Refugio Ruiz de Amador, 1860-1907',
    'Clotilde Amador de Terrazas, 1886-1945',
    'Antonio Terrazas, 1892-1919',
    'Corina Amador de Campbell, 1893-1924',
    'Emilia Amador de García, 1875-1942',
    'Jesus García, 1874-1922',
    'Francisco (Frank) Amador, 1883-1926',
    'Juan Amador, 1879-1909',
    'Julieta Amador de García, 1888-1949',
    'María Amador de Daguerre, 1886-1939',
    'Martin A. Amador, Jr., 1880-1889',
    'Miscellaneous, 1868-1944',
    'Personal Papers, 1892-1948'
]

def validate_series(value, series_values):
    """
    Validates if a series value is in the approved list and has the correct format.

    Parameters:
    - value (str): The series value to validate.
    - series_values (list): List of approved series names.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red' or 'yellow'), and a validation message.
    """
    # Clean the input value
    cleaned_value = value.strip() if isinstance(value, str) else ''
    print(f"Debug: Validating series value '{cleaned_value}'")

    # Check if the value is in the approved list
    if cleaned_value in series_values:
        print(f"Debug: Series name '{cleaned_value}' found in approved list. Validation passed.")
        return True, None, "Valid series name"

    # Check if a similar name exists but doesn't match exactly
    matching_series = [s for s in series_values if s.lower() == cleaned_value.lower()]
    if matching_series:
        print(f"Debug: Series name '{cleaned_value}' found but format is incorrect. Expected '{matching_series[0]}'.")
        return False, "red", f"Format error: Expected '{matching_series[0]}'"

    # If no match is found, highlight in yellow
    print(f"Debug: Series name '{cleaned_value}' not found in approved list. Highlighting in yellow.")
    return False, "yellow", "Series name not found in approved list"


def validate_box_folder(value, digital_identifier):
    """
    Validates that the BOX_FOLDER value matches the corresponding box and folder numbers
    extracted from the DIGITAL_IDENTIFIER.

    Parameters:
    - value (str): The BOX_FOLDER value to validate (e.g., '01_07').
    - digital_identifier (str): The DIGITAL_IDENTIFIER to extract box and folder numbers from (e.g., 'Ms0004_01_07_01.pdf').

    Returns:
    - (bool, str, str): Validation status, fill color ('red' or 'yellow'), and message.
    """
    # Check if both inputs are valid strings
    if not isinstance(value, str) or not isinstance(digital_identifier, str):
        return False, "red", "Invalid type: BOX_FOLDER or DIGITAL_IDENTIFIER is not a string"

    # Extract box and folder from DIGITAL_IDENTIFIER using regex
    match = re.match(r"^Ms\d{4}_(\d{2})_(\d{2})_\d{2}\.pdf$", digital_identifier)
    if not match:
        return False, "red", f"Invalid DIGITAL_IDENTIFIER format: '{digital_identifier}'"

    extracted_box, extracted_folder = match.groups()
    expected_box_folder = f"{extracted_box}_{extracted_folder}"

    # Compare with BOX_FOLDER value
    if value.strip() != expected_box_folder:
        return False, "red", f"BOX_FOLDER '{value.strip()}' does not match DIGITAL_IDENTIFIER box '{extracted_box}' and folder '{extracted_folder}'."

    # If everything matches
    return True, "", "BOX_FOLDER matches DIGITAL_IDENTIFIER"

    

def validate_collection_name(value, language="English"):
    """
    Validates the COLLECTION_NAME or ES..COLLECTION_NAME columns to ensure they match the expected value.

    Parameters:
    - value (str): The collection name to validate.
    - language (str): Language of the collection name ("English" or "Spanish").

    Returns:
    - (bool, str, str): Validation status, fill color ('red' for invalid), and a validation message.
    """
    # Define expected constants within the function
    EXPECTED_COLLECTION_NAME_EN = "Amador Family Correspondence, 1856-1949"
    EXPECTED_COLLECTION_NAME_ES = "Correspondencia de la familia Amador, 1856-1949"

    # Determine expected value based on language
    expected_value = (
        EXPECTED_COLLECTION_NAME_EN if language.lower() == "english" else EXPECTED_COLLECTION_NAME_ES
    )

    # Check if value matches the expected value
    if value.strip() == expected_value:
        return True, None, "Valid"
    else:
        return False, "red", f"Collection Name mismatch. Expected '{expected_value}', got '{value.strip()}'"



def is_valid_city_related(row, city_column, country_column, state_column, coord_column, language):
    city = str(row.get(city_column, '')).strip().lower() if pd.notna(row.get(city_column)) else ''
    country = row.get(country_column, '').strip() if pd.notna(row.get(country_column)) else ''
    state = row.get(state_column, '').strip() if pd.notna(row.get(state_column)) else ''
    coordinates = row.get(coord_column, '').strip() if pd.notna(row.get(coord_column)) else ''

    print(f"Debug: Validating city '{city}', country '{country}', state '{state}', coordinates '{coordinates}'")

    if not city or city == "no data":
        return True, "", "", "City data missing or marked as 'no data'"

    city_key = (city, language)
    expected_data = city_info.get(city_key)

    if not expected_data:
        return False, "yellow", city_column, f"City '{city}' not found in dataset for language '{language}'"

    if country and country != expected_data['country']:
        return False, "red", country_column, f"Country mismatch: Expected '{expected_data['country']}', found '{country}'"
    
    if state and state != expected_data['state']:
        return False, "red", state_column, f"State mismatch: Expected '{expected_data['state']}', found '{state}'"

    expected_coords = expected_data['coordinates']
    coord_sets = coordinates.split("[|]")

    matched_coords = [actual_coords.strip() for actual_coords in coord_sets if actual_coords.strip() == expected_coords]

    if not matched_coords:
        return False, "red", coord_column, f"No matching coordinates found for '{city}' with expected value '{expected_coords}'"

    if len(coord_sets) == 2 and coordinates != f"{coord_sets[0].strip()}[|]{coord_sets[1].strip()}":
        return False, "red", coord_column, "Coordinate format error: Incorrect '[|]' separator for dual coordinates."

    print("Debug: Location data matches expected values; validation passed.")
    return True, "", "", "Location data matches expected values"




# Adjust location validation rules to include the language parameter
location_validation_rules = {
    "SENDERS_CITY": lambda row: is_valid_city_related(row, 'SENDERS_CITY', 'SENDERS_COUNTRY', 'SENDERS_STATE', 'GEOLOC_SCITY', 'english'),
    "ES..SENDERS_CITY": lambda row: is_valid_city_related(row, 'ES..SENDERS_CITY', 'ES..SENDERS_COUNTRY', 'ES..SENDERS_STATE', 'ES..GEOLOC_SCITY', 'spanish'),
    "ADDRESSEES_CITY": lambda row: is_valid_city_related(row, 'ADDRESSEES_CITY', 'ADDRESSEES_COUNTRY', 'ADDRESSEES_STATE', 'GEOLOC_SCITY', 'english'),
    "ES..ADDRESSEES_CITY": lambda row: is_valid_city_related(row, 'ES..ADDRESSEES_CITY', 'ES..ADDRESSEES_COUNTRY', 'ES..ADDRESSEES_STATE', 'ES..GEOLOC_SCITY', 'spanish')
}




def validate_digital_identifier(value, previous_identifier=None):
    """
    Validate the DIGITAL_IDENTIFIER format for either 'Ms0004_XX_XX_XX.pdf' or 'Ms0071_XX_XX_XX.pdf' pattern.
    The format must follow 'Ms0004_<BoxNumber>_<FolderNumber>_<LetterNumber>.pdf' or 'Ms0071_<BoxNumber>_<FolderNumber>_<LetterNumber>.pdf'.
    The BoxNumber and FolderNumber should remain consistent in sequence, and
    LetterNumber should increment by 1 starting from '01' at row 2.

    Parameters:
    - value (str): The identifier to validate.
    - previous_identifier (tuple): The collection, box, folder, and letter numbers of the previous identifier.

    Returns:
    - (bool, tuple, str): Validation status, updated identifier, and message.
    """
    if not isinstance(value, str):
        return False, "red", "Invalid type: Expected a string"

    # Check format: Ms0004_XX_XX_XX.pdf or Ms0071_XX_XX_XX.pdf
    match = re.match(r"^(Ms0004|Ms0071)_(\d{2})_(\d{2})_(\d{2})\.pdf$", value)
    if not match:
        return False, "red", "Incorrect format. Expected 'Ms0004_XX_XX_XX.pdf' or 'Ms0071_XX_XX_XX.pdf' where XX are two-digit numbers"

    # Extract Collection, BoxNumber, FolderNumber, and LetterNumber
    collection, box_number, folder_number, letter_number = match.groups()

    # Convert letter_number to an integer for comparison
    current_letter_number = int(letter_number)

    # Check if this is the first identifier in the series (should be 01 at row 2)
    if previous_identifier is None:
        if current_letter_number != 1:
            return False, "red", "First letter number must start with 01"
        # Set initial tracking for the next row to increment from here
        return True, (collection, box_number, folder_number, current_letter_number), "Valid"

    # Unpack previous identifier
    prev_collection, prev_box, prev_folder, prev_letter = previous_identifier

    # Verify that Collection, BoxNumber, and FolderNumber match the previous identifier's values
    if collection != prev_collection or box_number != prev_box or folder_number != prev_folder:
        return False, "red", f"Box or folder number mismatch. Expected '{prev_box}_{prev_folder}' but got '{box_number}_{folder_number}'"

    # Calculate the next expected letter number, which should simply be previous letter + 1
    expected_letter_number = prev_letter + 1
    if current_letter_number != expected_letter_number:
        return False, "red", f"Letter number must increment sequentially. Expected {str(expected_letter_number).zfill(2)} but got {letter_number}"

    # If all checks pass, update previous identifier tracking and mark as valid
    return True, (collection, box_number, folder_number, current_letter_number), "Valid"



def is_valid_box_folder(value):
    if not isinstance(value, str):
        return False, None, "Box Folder value is not a string"
    if not re.match(r'^\d{2}_\d{2}$', value):
        return False, None, "Box Folder format is incorrect, expected 'XX_XX' with two digits before and after the underscore"
    return True, None, "Valid"




def is_valid_collection_name(value, language="English"):
    if not isinstance(value, str):
        return False, None, "Collection Name is not a string"
    
    expected_name = "Correspondencia de la familia Amador, 1856-1949" if language.lower() == "spanish" else "Amador Family Correspondence, 1856-1949"
    if value != expected_name:
        return False, None, f"Collection Name does not match expected value for {language}. Expected '{expected_name}' but got '{value}'"
    
    return True, None, "Valid"


def validate_collection_number(value):
    """
    Validates the collection number for both English and Spanish columns.
    The only valid values are 'Ms0004' and 'Ms0071'.

    Parameters:
    - value (str): The collection number to validate.

    Returns:
    - (bool, str, str): Tuple indicating if validation passed, the color for highlighting ('red'), 
      and an error message.
    """
    valid_values = {"Ms0004", "Ms0071"}

    # Clean and normalize the value
    cleaned_value = str(value).strip()

    if cleaned_value in valid_values:
        print(f"Validation successful: Collection number '{cleaned_value}' is valid.")
        return True, "", "Valid collection number"
    else:
        print(f"Failed validation: Collection number '{cleaned_value}' is invalid.")
        return False, "red", f"Invalid collection number: Expected one of {valid_values}, but got '{cleaned_value}'"





def is_valid_date(value):
    if pd.isna(value):
        return True, None, "Valid (empty)"
    try:
        pd.to_datetime(value, format='%Y-%m-%d', errors='raise')
        return True, None, "Valid"
    except (ValueError, TypeError):
        pass
    try:
        pd.to_datetime(value, format='%Y-%m', errors='raise')
        return True, None, "Valid"
    except (ValueError, TypeError):
        return False, None, "Date format is invalid. Expected 'YYYY-MM-DD' or 'YYYY-MM'"




def is_valid_year(value):
    if not isinstance(value, int):
        return False, None, "Year is not an integer"
    if not (1000 <= value <= 9999):
        return False, None, "Year is out of valid range (1000-9999)"
    return True, None, "Valid"



def validate_name_field(value, authorized_names):
    """
    Validates if a name in the FROM or TO fields matches an entry in the authorized names dataset.
    If the name is missing from the dataset entirely, it highlights the cell in yellow.
    If the name exists in the dataset but is not in the correct format, it highlights the cell in red.

    Parameters:
    - value (str): The name to validate.
    - authorized_names (set): Set of authorized names loaded from the dataset.

    Returns:
    - (bool, str, str): Validation status, fill color ('red', 'yellow', or None), and message.
    """
    known_unknown_values = {
        "FROM" : {"Unknown sender"},
        "ES..FROM": {"Remitente desconocido"},
        "TO" : {"Unknown recipient"},
        "ES..TO": {"Destinatario desconocido"}

    }

    # Clean up the input value
    cleaned_value = value.strip()
    
    print(f"Debug: Starting search for '{cleaned_value}' in authorized names...")

    # Check if the value is in the known "unknown" set
    if cleaned_value in known_unknown_values:
        print(f"Debug: '{cleaned_value}' recognized as a known unknown value. Validation passed.")
        return True, None, "Valid (unknown value)"
    
    # Check if cleaned_value matches an authorized name exactly
    if cleaned_value in authorized_names:
        print(f"Debug: '{cleaned_value}' found in authorized names with correct format. Validation passed.")
        return True, None, "Valid"
    
    # Check if a similar name exists but does not match exactly
    matching_names = [name for name in authorized_names if name.lower() == cleaned_value.lower()]
    if matching_names:
        print(f"Debug: '{cleaned_value}' found in authorized names but format is incorrect. Expected '{matching_names[0]}'.")
        return False, "red", f"Format error: Expected '{matching_names[0]}'"
    
    # Name does not exist in the authorized names dataset at all
    print(f"Debug: '{cleaned_value}' not found in authorized names dataset.")
    return False, "yellow", "Name not found in dataset"


def is_valid_subject_lcsh(value, approved_subjects, language="english"):
    if not isinstance(value, str):
        return False, None, "Invalid type: Expected a string"

    terms = [term.strip() for term in value.split("[|]")]
    print(f"Debug: Starting validation for SUBJECT_LCSH terms '{terms}' in language '{language}'...")

    invalid_terms = []
    for term in terms:
        if term in approved_subjects[language]:
            print(f"Debug: Term '{term}' found in vocabulary for '{language}'. Validation passed.")
        else:
            print(f"Debug: Term '{term}' NOT found in vocabulary for '{language}'. Validation failed.")
            invalid_terms.append(term)

    if invalid_terms:
        return False, "yellow", f"Terms not found in vocabulary: {invalid_terms}"
    return True, None, "Valid"



def load_authorized_names(names_dataset_path):
    names_data = pd.read_excel(names_dataset_path, usecols=[0], header=None)
    authorized_names = set(names_data[0].dropna().str.strip())
    print("Debug: Authorized names loaded from dataset:", authorized_names)
    return authorized_names

# Column validation rules
column_validation_rules = {
    "DIGITAL_IDENTIFIER": validate_digital_identifier,
    "ES..DIGITAL_IDENTIFIER": validate_digital_identifier,
    "BOX_FOLDER": is_valid_box_folder,
    "ES..BOX_FOLDER": is_valid_box_folder,
    "COLLECTION_NAME": lambda x: is_valid_collection_name(x, language="English"),
    "ES..COLLECTION_NAME": lambda x: is_valid_collection_name(x, language="Spanish"),
    "DATE": is_valid_date,
    "ES..DATE": is_valid_date,
    "YEAR": is_valid_year,
    "ES..YEAR": is_valid_year,
    "SUBJECT_LCSH": lambda x: is_valid_subject_lcsh(x, approved_subjects, language="english"),
    "ES..SUBJECT_LCSH" : lambda x: is_valid_subject_lcsh(x, approved_subjects, language="spanish"),
    "FROM": lambda x: validate_name_field(x, authorized_names),
    "ES..FROM": lambda x: validate_name_field(x, authorized_names),
    "TO": lambda x: validate_name_field(x, authorized_names),
    "ES..TO": lambda x: validate_name_field(x, authorized_names),
    # Add validation rules to the main column validation structure
    "SERIES" : validate_series,
    "ES..SERIES" : validate_series
}

location_validation_rules = {
    "SENDERS_CITY": lambda row: is_valid_city_related(row, 'SENDERS_CITY', 'SENDERS_COUNTRY', 'SENDERS_STATE', 'GEOLOC_SCITY', 'english'),
    "ES..SENDERS_CITY": lambda row: is_valid_city_related(row, 'ES..SENDERS_CITY', 'ES..SENDERS_COUNTRY', 'ES..SENDERS_STATE', 'ES..GEOLOC_SCITY', 'spanish'),
    "ADDRESSEES_CITY": lambda row: is_valid_city_related(row, 'ADDRESSEES_CITY', 'ADDRESSEES_COUNTRY', 'ADDRESSEES_STATE', 'GEOLOC_SCITY', 'english'),
    "ES..ADDRESSEES_CITY": lambda row: is_valid_city_related(row, 'ES..ADDRESSEES_CITY', 'ES..ADDRESSEES_COUNTRY', 'ES..ADDRESSEES_STATE', 'ES..GEOLOC_SCITY', 'spanish')
}



approved_subjects = load_approved_subjects("SUBJECT_LCSH.xlsx")
city_info = load_city_data("Maybeee.xlsx")
authorized_names = load_authorized_names("CVPeople.xlsx")
print("Loaded authorized names:", authorized_names)

# The `verify_file` function and main script setup remain the same, using `column_validation_rules`.

def verify_file(input_file, output_file):
    # Load the Excel file and worksheet
    wb = load_workbook(input_file)
    ws = wb["OA_Descriptive metadata"]
    df = pd.read_excel(input_file, sheet_name="OA_Descriptive metadata")

    # Initialize previous identifiers separately for each column
    previous_identifier_digital = None  # For DIGITAL_IDENTIFIER column
    previous_identifier_es_digital = None  # For ES..DIGITAL_IDENTIFIER column

    for idx, row in df.iterrows():
        # Validate non-location columns
        for col_name, validation_func in column_validation_rules.items():
            if col_name in df.columns:
                value = row[col_name]
                try:
                    # Special handling for DIGITAL_IDENTIFIER to track sequence
                    if col_name == "DIGITAL_IDENTIFIER":
                        is_valid, result, message = validate_digital_identifier(value, previous_identifier_digital)
                        if is_valid:
                            previous_identifier_digital = result  # Update tracking for next row
                            print(f"Validation successful: {col_name} at row {idx + 2}")
                        else:
                            # Apply the correct color highlight based on the validation result
                            col_idx = df.columns.get_loc(col_name) + 1
                            ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if result == "red" else highlight_fill_yellow
                            print(f"Failed validation: {col_name} at row {idx + 2} - Reason: {message}")
                    
                    elif col_name == "ES..DIGITAL_IDENTIFIER":
                        is_valid, result, message = validate_digital_identifier(value, previous_identifier_es_digital)
                        if is_valid:
                            previous_identifier_es_digital = result  # Update tracking for next row
                            print(f"Validation successful: {col_name} at row {idx + 2}")
                        else:
                            col_idx = df.columns.get_loc(col_name) + 1
                            ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if result == "red" else highlight_fill_yellow
                            print(f"Failed validation: {col_name} at row {idx + 2} - Reason: {message}")

                    else:
                        # General validation for non-DIGITAL_IDENTIFIER columns
                        is_valid, color, message = validation_func(value)
                        if is_valid:
                            print(f"Validation successful: {col_name} at row {idx + 2}")
                        else:
                            col_idx = df.columns.get_loc(col_name) + 1
                            ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                            print(f"Failed validation: {col_name} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col_name} at row {idx + 2}: {e}")
        
        # Validate SERIES and ES..SERIES columns
        if "SERIES" in df.columns:
            value = row["SERIES"]
            try:
                is_valid, color, message = validate_series(value, series_values)
                if is_valid:
                    print(f"Validation successful: SERIES at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("SERIES") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: SERIES at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating SERIES at row {idx + 2}: {e}")

        if "ES..SERIES" in df.columns:
            value = row["ES..SERIES"]
            try:
                is_valid, color, message = validate_series(value, series_values)
                if is_valid:
                    print(f"Validation successful: ES..SERIES at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("ES..SERIES") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: ES..SERIES at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating ES..SERIES at row {idx + 2}: {e}")

        # Validate COLLECTION_NAME and ES..COLLECTION_NAME
        if "COLLECTION_NAME" in df.columns:
            value = row["COLLECTION_NAME"]
            try:
                is_valid, color, message = validate_collection_name(value, "English")
                if is_valid:
                    print(f"Validation successful: COLLECTION_NAME at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("COLLECTION_NAME") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: COLLECTION_NAME at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating COLLECTION_NAME at row {idx + 2}: {e}")
        
        if "ES..COLLECTION_NAME" in df.columns:
            value = row["ES..COLLECTION_NAME"]
            try:
                is_valid, color, message = validate_collection_name(value, "Spanish")
                if is_valid:
                    print(f"Validation successful: ES..COLLECTION_NAME at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("ES..COLLECTION_NAME") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: ES..COLLECTION_NAME at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating ES..COLLECTION_NAME at row {idx + 2}: {e}")
                
        # Validate location-related columns
        for loc_col_name, location_func in location_validation_rules.items():
            if loc_col_name in df.columns:
                try:
                    is_valid, color, highlight_col, message = location_func(row)
                    if is_valid:
                        print(f"Location validation successful: {loc_col_name} at row {idx + 2}")
                    else:
                        # Highlight the specific failing column
                        col_idx = df.columns.get_loc(highlight_col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                        print(f"Failed location validation: {loc_col_name} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error in location validation for {loc_col_name} at row {idx + 2}: {e}")

    # Save the workbook after validation and highlighting
    wb.save(output_file)
    print(f"Verification completed. Output saved as {output_file}")








# Argument parser to take file input and generate verified output
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Verify an Excel file against validation rules.")
    parser.add_argument("file_name", help="The name of the Excel file to verify")
    args = parser.parse_args()
    
    # Generate output file name with "Verified_" prefix
    input_file = args.file_name
    output_file = f"Verified_{os.path.basename(input_file)}"
    
    # Run verification
    verify_file(input_file, output_file)

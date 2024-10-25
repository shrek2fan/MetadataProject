import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def compare_files(original_file, cleaned_file, output_file):
    original_sheets = pd.read_excel(original_file, sheet_name=None)
    cleaned_sheets = pd.read_excel(cleaned_file, sheet_name=None)

    # Debugging: Print loaded sheet names
    print(f"Original file sheets: {original_sheets.keys()}")
    print(f"Cleaned file sheets: {cleaned_sheets.keys()}")
    
    # Dictionary to hold comparison DataFrames
    comparison_sheets = {}
    
    for sheet_name, original_df in original_sheets.items():
        if sheet_name not in cleaned_sheets:
            print(f"Skipping sheet: {sheet_name} (not in both files)")
            continue
        
        cleaned_df = cleaned_sheets[sheet_name]
        print(f"Comparing sheet: {sheet_name}")
        
        # Ensure both dataframes have the same columns
        original_columns = set(original_df.columns)
        cleaned_columns = set(cleaned_df.columns)
        
        if original_columns == cleaned_columns:
            # Create a new DataFrame for combining original and cleaned data
            combined_df = pd.DataFrame()
            
            for column in original_df.columns:
                combined_column = pd.concat([
                    original_df[column], 
                    pd.Series([''] * 5),  # 5 blank rows
                    cleaned_df[column].reset_index(drop=True)
                ], ignore_index=True)
                
                combined_df[column] = combined_column
            
            # Debugging: Check if the combined DataFrame is being created
            print(f"Combined DataFrame for {sheet_name}:")
            print(combined_df.head())

            comparison_sheets[sheet_name] = combined_df
        else:
            print(f"Warning: Columns in sheet {sheet_name} do not match!")
            print(f"Original columns: {original_columns}")
            print(f"Cleaned columns: {cleaned_columns}")
            print(f"Difference in columns: {original_columns.symmetric_difference(cleaned_columns)}")
    
    # Save the comparison if any sheets were created
    if comparison_sheets:
        save_comparison(comparison_sheets, output_file)
    else:
        print("No comparison sheets were generated!")


def save_comparison(comparison_sheets, output_file):
    # Create a new Excel workbook
    wb = Workbook()
    
    # Remove default empty sheet
    if 'Sheet' in wb.sheetnames and not comparison_sheets:
        wb.remove(wb['Sheet'])
    
    # Debugging output
    print(f"Saving comparison file: {output_file}")
    
    # Add comparison sheets
    if comparison_sheets:
        for sheet_name, comparison_df in comparison_sheets.items():
            print(f"Adding sheet: {sheet_name} to comparison file")
            ws = wb.create_sheet(title=sheet_name)
            
            # Write DataFrame to rows, ensuring original and transformed values are in the same column
            for row in dataframe_to_rows(comparison_df, index=False, header=True):
                ws.append(row)
    else:
        # If no comparisons found, add a default sheet to avoid error
        ws = wb.create_sheet(title="No Comparisons")
        ws.append(["No differences were found."])
        print("No differences found, adding empty sheet.")
    
    # Save the workbook to the output file
    try:
        wb.save(output_file)
        print(f"Comparison saved successfully to {output_file}")
    except Exception as e:
        print(f"Error saving file: {e}")



if __name__ == "__main__":
    print("Calling compare_files function...")  # <-- Add this print statement
    import argparse
    parser = argparse.ArgumentParser(description="Compare two Excel files.")
    parser.add_argument("original_file", help="Path to the original Excel file")
    parser.add_argument("cleaned_file", help="Path to the cleaned Excel file")
    parser.add_argument("output_file", help="Path to save the comparison output file")
    args = parser.parse_args()
    
    compare_files(args.original_file, args.cleaned_file, args.output_file)
import openpyxl
import argparse
from openpyxl.styles import PatternFill

# Define red fill for highlighting invalid cells
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Load the Excel file
def load_excel(file_path):
    return openpyxl.load_workbook(file_path)

# Save the Excel file after processing
def save_excel(workbook, file_path):
    workbook.save(file_path)

# Example function to check if the value is valid
def check_digital_identifier(cell_value):
    # Custom logic to check if the digital identifier is valid
    if isinstance(cell_value, str) and cell_value.endswith('.pdf') and cell_value.startswith('Ms0004'):
        return True
    return False

def check_title(cell_value):
    # Logic for titles (just an example)
    if isinstance(cell_value, str) and cell_value.strip():
        return True
    return False

# Add more validation functions as needed for each column
# ...

# Process each sheet and highlight invalid cells
def highlight_invalid_cells(file_path, output_file):
    workbook = load_excel(file_path)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Processing sheet: {sheet_name}")
        
        # Check for specific columns and highlight invalid cells
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if sheet.cell(row=1, column=cell.column).value == "DIGITAL_IDENTIFIER":
                    if not check_digital_identifier(cell.value):
                        cell.fill = red_fill
                    else:
                        cell.fill = green_fill  # Success - Highlight green
                elif sheet.cell(row=1, column=cell.column).value == "TITLE":
                    if not check_title(cell.value):
                        cell.fill = red_fill
                    else:
                        cell.fill = green_fill
                # Add checks for other columns as needed
                # elif sheet.cell(row=1, column=cell.column).value == "OTHER_COLUMN":
                #     if not check_other_column(cell.value):
                #         cell.fill = red_fill
                #     else:
                #         cell.fill = green_fill

    # Save the workbook with the highlighted cells
    save_excel(workbook, output_file)
    print(f"Highlighting complete. File saved as: {output_file}")

# Main function to run the script
if __name__ == "__main__":
    # Argument parsing to accept the file name as input
    parser = argparse.ArgumentParser(description="Highlight invalid cells in an Excel file.")
    parser.add_argument("file_name", help="The name of the Excel file to process (e.g., Test7.xlsx)")
    args = parser.parse_args()

    input_file = args.file_name  # Input Excel file passed as an argument
    output_file = f"Highlighted_{input_file}"  # Output file will prepend 'Highlighted_' to the filename

    highlight_invalid_cells(input_file, output_file)

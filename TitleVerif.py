import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import unidecode
import sys
import os

# Define the green fill for highlighting successfully processed cells
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Function to correct accented characters if they were lost or malformed
def correct_accents(text):
    return unidecode.unidecode_expect_ascii(text)  # This preserves accents where needed

# Function to format the title according to sentence logic, ensuring proper accents
def format_title_sentence_logic(title):
    title = correct_accents(title.strip())  # Ensure accents and strip whitespace
    
    # Split title into words
    words = title.split()
    
    # Capitalize only the first word and proper names
    if words:
        words[0] = words[0].capitalize()  # Always capitalize the first word
        for i in range(1, len(words)):
            if is_proper_noun(words[i]):
                words[i] = words[i].capitalize()  # Capitalize proper names
            else:
                words[i] = words[i].lower()  # Lowercase other words (like "for", "a", etc.)
    
    # Join the words back into a title
    return " ".join(words)

# Function to detect proper nouns (for simplicity, detecting words starting with uppercase letters)
def is_proper_noun(word):
    return word.istitle()  # A proper noun starts with an uppercase letter

# Main function to process the input file
def process_file(input_file):
    print(f"Processing file: {input_file}")
    
    # Generate the output filename based on the input file name
    base_filename = os.path.splitext(os.path.basename(input_file))[0]
    output_file = f"Transformed_{base_filename}.xlsx"

    # Load the workbook
    wb = load_workbook(input_file)
    sheet = wb['OA_Descriptive metadata']  # Process only the relevant sheet

    title_column_index = 5  # Assuming TITLE is column 5

    # Loop through rows in the 'TITLE' column
    for row in range(2, sheet.max_row + 1):
        title_cell = sheet.cell(row=row, column=title_column_index)
        if title_cell.value:
            original_title = title_cell.value  # Debugging: capture the original value
            
            # Apply sentence logic formatting and accent correction
            formatted_title = format_title_sentence_logic(original_title)

            # Debugging: print original and transformed value
            print(f"Row {row}: TITLE before: '{original_title}', after: '{formatted_title}'")

            # Update the cell value with the newly formatted title
            title_cell.value = formatted_title
            title_cell.fill = green_fill  # Highlight processed cell

    # Save the transformed workbook with a new filename
    wb.save(output_file)
    print(f"File processed and saved as: {output_file}")

# Main execution: check for command-line arguments
if __name__ == "__main__":
    # Check if an input file was provided as an argument
    if len(sys.argv) < 2:
        print("Error: No input file specified.")
        print("Usage: python TitleVerif.py <input_file.xlsx>")
        sys.exit(1)
    
    # Get the input file from command-line argument
    input_file = sys.argv[1]
    
    # Process the input file
    process_file(input_file)

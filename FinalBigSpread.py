import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import argparse
import os
import datetime 
import logging
import sys



# Add the PrintToLogger code block here
# Redirect print statements to logging and console
logging.basicConfig(
    filename='validation_log.txt',
    level=logging.DEBUG,
    format='%(asctime)s - %(message)s'
)

class PrintToLogger:
    def __init__(self):
        self.console = sys.stdout

    def write(self, message):
        if message.strip():  # Avoid empty log entries
            logging.info(message.strip())
            self.console.write(message)

    def flush(self):  # Required for Python's `sys.stdout`
        pass

sys.stdout = PrintToLogger()



# Define fill styles for highlighting mistakes
highlight_fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
highlight_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Load the approved SUBJECT_LCSH vocabulary with separate English and Spanish terms
def load_approved_subjects(vocabulary_file):
    approved_df = pd.read_excel(vocabulary_file, usecols=[0, 1])  # Load only columns A and B
    approved_subjects = {
        "spanish": set(approved_df.iloc[:, 0].dropna().str.strip()),  # Spanish terms from column A
        "english": set(approved_df.iloc[:, 1].dropna().str.strip()),  # English terms from column B
    }
    print("Debug: Approved subjects loaded:", approved_subjects)  # Log loaded terms for confirmation
    return approved_subjects


# Load the city dataset into a dictionary structure with safe handling for non-string values
def load_city_data(city_dataset_path):
    city_data = pd.read_excel(city_dataset_path)

    def safe_strip(value):
        return str(value).strip() if isinstance(value, str) else ''

    city_info = {}
    for _, row in city_data.iterrows():
        city_info[(safe_strip(row['ES_City']).lower(), 'spanish')] = {
            'country': safe_strip(row['ES_Country']),
            'state': safe_strip(row['ES_State']),
            'coordinates': safe_strip(row["CITIES' LAT_LONG COORDINATES"])
        }
        city_info[(safe_strip(row['EN_City']).lower(), 'english')] = {
            'country': safe_strip(row['EN_Country']),
            'state': safe_strip(row['EN_State']),
            'coordinates': safe_strip(row["CITIES' LAT_LONG COORDINATES"])
        }
    return city_info


# List of valid series names
series_values = [
    'Martin Amador, 1856-1904',
    'Refugio Ruiz de Amador, 1860-1907',
    'Clotilde Amador de Terrazas, 1886-1945',
    'Antonio Terrazas, 1892-1919',
    'Corina Amador de Campbell, 1893-1924',
    'Emilia Amador de García, 1875-1942',
    'Jesus García, 1874-1922',
    'Francisco (Frank) Amador, 1883-1926',
    'Juan Amador, 1879-1909',
    'Julieta Amador de García, 1888-1949',
    'María Amador de Daguerre, 1886-1939',
    'Martin A. Amador, Jr., 1880-1889',
    'Miscellaneous, 1868-1944',
    'Personal Papers, 1892-1948'
]

# RELATIONSHIP 1 and RELATIONSHIP 2 mapping
relationship_mapping = {
    "Familia": {
        "Cuñado y Cuñada", "Cuñados", "Cuñadas", "Esposo y Esposa", "Prometidos", "Hermanas",
        "Hermanos", "Madre e Hija", "Madre e Hijo", "Padre e Hijo", "Padre e Hija", "Padres e Hijo",
        "Padres e Hija", "Padre e Hijos", "Primos", "Primas", "Tia y Sobrina", "Tia y Sobrino",
        "Tio y Sobrina", "Tio y Sobrino", "Tio, Tia y Sobrina", "Suegra y Yerno", "Suegra y Nuera",
        "Suegro y Yerno", "Abuela y Nieto", "Abuelo y Nieto", "Abuela y Nieta"
    },
    "Amigos": { "Compañeros de Escuela" },
    "Conocidos": set(),
    "Maestro y Estudiante": set(),
    "Compadrazgo": set(),
    "Comunidad": set(),
    "Iglesia Católica y Feligréses": { "Capellán y Feligrés" },
    "Socios Comerciales": { " Abogado y Cliente", "Propietario y Prospecto", "Vendedor y Cliente", "Médico y Paciente", "Propietario e Inquilino "}
}



# English translation for RELATIONSHIP 1 and RELATIONSHIP 2
relationship_mapping_english = {
    "Family": {
        "Brother-in-law and Sister-in-law", "Brothers-in-law", "Sisters-in-law", "Husband and Wife",
        "Fiancés", "Sisters", "Siblings", "Mother and Daughter", "Mother and Son", "Father and Son",
        "Father and Daughter", "Parents and Son", "Parents and Daughter", "Father and Children",
        "Cousins", "Aunt and Niece", "Aunt and Nephew", "Uncle and Niece", "Uncle and Nephew",
        "Uncle, Aunt and Niece", "Mother-in-law and Son-in-law", "Mother-in-Law and Daughter-in-Law",
        "Father-in-law and Son-in-law", "Grandmother and Grandson", "Grandfather and Grandson",
        "Grandmother and Granddaughter"
    },
    "Friends": { "Schoolmates" },
    "Acquaintances": set(),
    "Teacher and Student": set(),
    "God Parenthood": set(),
    "Community Organization and Citizen": set(),
    "Catholic Church and Churchgoers": { "Chaplain and Churchgoer" },
    "Business Partners": { "Lawyer and Client", "Proprietor and Prospect", "Seller and Client",  "Doctor and Patient", "Landlord and Tenant" }
}





def validate_series(value, series_values):
    """
    Validates if a series value is in the approved list and has the correct format.

    Parameters:
    - value (str): The series value to validate.
    - series_values (list): List of approved series names.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red' or 'yellow'), and a validation message.
    """
    # Clean the input value
    cleaned_value = value.strip() if isinstance(value, str) else ''
    print(f"Debug: Validating series value '{cleaned_value}'")

    # Check if the value is in the approved list
    if cleaned_value in series_values:
        print(f"Debug: Series name '{cleaned_value}' found in approved list. Validation passed.")
        return True, None, "Valid series name"

    # Check if a similar name exists but doesn't match exactly
    matching_series = [s for s in series_values if s.lower() == cleaned_value.lower()]
    if matching_series:
        print(f"Debug: Series name '{cleaned_value}' found but format is incorrect. Expected '{matching_series[0]}'.")
        return False, "red", f"Format error: Expected '{matching_series[0]}'"

    # If no match is found, highlight in yellow
    print(f"Debug: Series name '{cleaned_value}' not found in approved list. Highlighting in yellow.")
    return False, "yellow", "Series name not found in approved list"



def validate_relationships(rel1_value, rel2_value, mapping, lang, column_name_rel1, column_name_rel2):
    """
    Validates RELATIONSHIP 1 and RELATIONSHIP 2 columns with support for multiple terms in RELATIONSHIP 1.

    Parameters:
    - rel1_value (str): The value in the RELATIONSHIP 1 column.
    - rel2_value (str): The value in the RELATIONSHIP 2 column.
    - mapping (dict): The mapping of RELATIONSHIP 1 terms to valid RELATIONSHIP 2 terms.
    - lang (str): Language of the validation ("English" or "Spanish").
    - column_name_rel1 (str): Name of the RELATIONSHIP 1 column being validated.
    - column_name_rel2 (str): Name of the RELATIONSHIP 2 column being validated.

    Returns:
    - (bool, str, str): Tuple indicating if validation passed, the color for highlighting ('red' or 'yellow'),
      and an error message.
    """
    try:
        # Clean and strip values
        rel1_cleaned = str(rel1_value).strip() if pd.notna(rel1_value) else ""
        rel2_cleaned = str(rel2_value).strip() if pd.notna(rel2_value) else ""

        # Skip validation for empty RELATIONSHIP 1 and RELATIONSHIP 2
        if not rel1_cleaned and not rel2_cleaned:
            return True, "", "No validation needed for empty relationships."

        # Split RELATIONSHIP 1 terms if they are separated by [|]
        rel1_terms = [term.strip() for term in rel1_cleaned.split("[|]")]

        # Check for duplicates in RELATIONSHIP 1 terms
        if len(rel1_terms) != len(set(rel1_terms)):
            return False, "red", f"Duplicate terms found in {column_name_rel1}: '{rel1_cleaned}'"

        # Validate each term in RELATIONSHIP 1
        invalid_terms = [term for term in rel1_terms if term not in mapping]
        if invalid_terms:
            return False, "red", f"Invalid {column_name_rel1} terms: {', '.join(invalid_terms)} in {lang}"

        # Validate RELATIONSHIP 2, if provided
        if rel2_cleaned:
            valid_rel2 = set()
            for term in rel1_terms:
                valid_rel2.update(mapping.get(term, set()))
            if rel2_cleaned not in valid_rel2:
                return False, "red", f"Invalid {column_name_rel2} value: '{rel2_cleaned}' for {column_name_rel1}: '{rel1_cleaned}' in {lang}"

        # If all checks pass
        return True, "", "Valid relationship values"
    except Exception as e:
        return False, "red", f"Error validating relationships in {lang}: {str(e)}"

# Function updated. It now supports multiple terms in RELATIONSHIP1 and validates them against mapping. Let me know if further modifications are needed.







def validate_box_folder(value, digital_identifier):
    """
    Validates that the BOX_FOLDER value matches the corresponding box and folder numbers
    extracted from the DIGITAL_IDENTIFIER.

    Parameters:
    - value (str): The BOX_FOLDER value to validate (e.g., '01_07').
    - digital_identifier (str): The DIGITAL_IDENTIFIER to extract box and folder numbers from (e.g., 'Ms0004_01_07_01.pdf').

    Returns:
    - (bool, str, str): Validation status, fill color ('red' or 'yellow'), and message.
    """
    # Check if both inputs are valid strings
    if not isinstance(value, str) or not isinstance(digital_identifier, str):
        return False, "red", "Invalid type: BOX_FOLDER or DIGITAL_IDENTIFIER is not a string"

    # Extract box and folder from DIGITAL_IDENTIFIER using regex
    match = re.match(r"^Ms\d{4}_(\d{2})_(\d{2})_\d{2}\.pdf$", digital_identifier)
    if not match:
        return False, "red", f"Invalid DIGITAL_IDENTIFIER format: '{digital_identifier}'"

    extracted_box, extracted_folder = match.groups()
    expected_box_folder = f"{extracted_box}_{extracted_folder}"

    # Compare with BOX_FOLDER value
    if value.strip() != expected_box_folder:
        return False, "red", f"BOX_FOLDER '{value.strip()}' does not match DIGITAL_IDENTIFIER box '{extracted_box}' and folder '{extracted_folder}'."

    # If everything matches
    return True, "", "BOX_FOLDER matches DIGITAL_IDENTIFIER"

    

def validate_collection_name(value, language="English"):
    """
    Validates the COLLECTION_NAME or ES..COLLECTION_NAME columns to ensure they match the expected value.

    Parameters:
    - value (str): The collection name to validate.
    - language (str): Language of the collection name ("English" or "Spanish").

    Returns:
    - (bool, str, str): Validation status, fill color ('red' for invalid), and a validation message.
    """
    # Define expected constants within the function
    EXPECTED_COLLECTION_NAME_EN = "Amador Family Correspondence, 1856-1949"
    EXPECTED_COLLECTION_NAME_ES = "Correspondencia de la familia Amador, 1856-1949"

    # Determine expected value based on language
    expected_value = (
        EXPECTED_COLLECTION_NAME_EN if language.lower() == "english" else EXPECTED_COLLECTION_NAME_ES
    )

    # Check if value matches the expected value
    if value.strip() == expected_value:
        return True, None, "Valid"
    else:
        return False, "red", f"Collection Name mismatch. Expected '{expected_value}', got '{value.strip()}'"



def is_valid_city_related(row, city_column, country_column, state_column, coord_column, language):
    city = str(row.get(city_column, '')).strip().lower() if pd.notna(row.get(city_column)) else ''
    country = row.get(country_column, '').strip() if pd.notna(row.get(country_column)) else ''
    state = row.get(state_column, '').strip() if pd.notna(row.get(state_column)) else ''
    coordinates = row.get(coord_column, '').strip() if pd.notna(row.get(coord_column)) else ''

    print(f"Debug: Validating city '{city}', country '{country}', state '{state}', coordinates '{coordinates}'")

    if not city or city == "no data":
        return True, "", "", "City data missing or marked as 'no data'"

    city_key = (city, language)
    expected_data = city_info.get(city_key)

    if not expected_data:
        return False, "yellow", city_column, f"City '{city}' not found in dataset for language '{language}'"

    if country and country != expected_data['country']:
        return False, "red", country_column, f"Country mismatch: Expected '{expected_data['country']}', found '{country}'"
    
    if state and state != expected_data['state']:
        return False, "red", state_column, f"State mismatch: Expected '{expected_data['state']}', found '{state}'"

    expected_coords = expected_data['coordinates']
    coord_sets = coordinates.split("[|]")

    matched_coords = [actual_coords.strip() for actual_coords in coord_sets if actual_coords.strip() == expected_coords]

    if not matched_coords:
        return False, "red", coord_column, f"No matching coordinates found for '{city}' with expected value '{expected_coords}'"

    if len(coord_sets) == 2 and coordinates != f"{coord_sets[0].strip()}[|]{coord_sets[1].strip()}":
        return False, "red", coord_column, "Coordinate format error: Incorrect '[|]' separator for dual coordinates."

    print("Debug: Location data matches expected values; validation passed.")
    return True, "", "", "Location data matches expected values"




# Adjust location validation rules to include the language parameter
location_validation_rules = {
    "SENDERS_CITY": lambda row: is_valid_city_related(row, 'SENDERS_CITY', 'SENDERS_COUNTRY', 'SENDERS_STATE', 'GEOLOC_SCITY', 'english'),
    "ES..SENDERS_CITY": lambda row: is_valid_city_related(row, 'ES..SENDERS_CITY', 'ES..SENDERS_COUNTRY', 'ES..SENDERS_STATE', 'ES..GEOLOC_SCITY', 'spanish'),
    "ADDRESSEES_CITY": lambda row: is_valid_city_related(row, 'ADDRESSEES_CITY', 'ADDRESSEES_COUNTRY', 'ADDRESSEES_STATE', 'GEOLOC_SCITY', 'english'),
    "ES..ADDRESSEES_CITY": lambda row: is_valid_city_related(row, 'ES..ADDRESSEES_CITY', 'ES..ADDRESSEES_COUNTRY', 'ES..ADDRESSEES_STATE', 'ES..GEOLOC_SCITY', 'spanish')
}




def validate_digital_identifier(value, previous_identifier=None):
    """
    Validate the DIGITAL_IDENTIFIER format for either 'Ms0004_XX_XX_XX.pdf' or 'Ms0071_XX_XX_XX.pdf' pattern.
    The format must follow 'Ms0004_<BoxNumber>_<FolderNumber>_<LetterNumber>.pdf' or 'Ms0071_<BoxNumber>_<FolderNumber>_<LetterNumber>.pdf'.
    The BoxNumber and FolderNumber should remain consistent in sequence, and
    LetterNumber should increment by 1 starting from '01' at row 2.

    Parameters:
    - value (str): The identifier to validate.
    - previous_identifier (tuple): The collection, box, folder, and letter numbers of the previous identifier.

    Returns:
    - (bool, tuple, str): Validation status, updated identifier, and message.
    """
    if not isinstance(value, str):
        return False, "red", "Invalid type: Expected a string"

    # Check format: Ms0004_XX_XX_XX.pdf or Ms0071_XX_XX_XX.pdf
    match = re.match(r"^(Ms0004|Ms0071)_(\d{2})_(\d{2})_(\d{2})\.pdf$", value)
    if not match:
        return False, "red", "Incorrect format. Expected 'Ms0004_XX_XX_XX.pdf' or 'Ms0071_XX_XX_XX.pdf' where XX are two-digit numbers"

    # Extract Collection, BoxNumber, FolderNumber, and LetterNumber
    collection, box_number, folder_number, letter_number = match.groups()

    # Convert letter_number to an integer for comparison
    current_letter_number = int(letter_number)

    # Check if this is the first identifier in the series (should be 01 at row 2)
    if previous_identifier is None:
        if current_letter_number != 1:
            return False, "red", "First letter number must start with 01"
        # Set initial tracking for the next row to increment from here
        return True, (collection, box_number, folder_number, current_letter_number), "Valid"

    # Unpack previous identifier
    prev_collection, prev_box, prev_folder, prev_letter = previous_identifier

    # Verify that Collection, BoxNumber, and FolderNumber match the previous identifier's values
    if collection != prev_collection or box_number != prev_box or folder_number != prev_folder:
        return False, "red", f"Box or folder number mismatch. Expected '{prev_box}_{prev_folder}' but got '{box_number}_{folder_number}'"

    # Calculate the next expected letter number, which should simply be previous letter + 1
    expected_letter_number = prev_letter + 1
    if current_letter_number != expected_letter_number:
        return False, "red", f"Letter number must increment sequentially. Expected {str(expected_letter_number).zfill(2)} but got {letter_number}"

    # If all checks pass, update previous identifier tracking and mark as valid
    return True, (collection, box_number, folder_number, current_letter_number), "Valid"



def is_valid_box_folder(value):
    if not isinstance(value, str):
        return False, None, "Box Folder value is not a string"
    if not re.match(r'^\d{2}_\d{2}$', value):
        return False, None, "Box Folder format is incorrect, expected 'XX_XX' with two digits before and after the underscore"
    return True, None, "Valid"




def is_valid_collection_name(value, language="English"):
    if not isinstance(value, str):
        return False, None, "Collection Name is not a string"
    
    expected_name = "Correspondencia de la familia Amador, 1856-1949" if language.lower() == "spanish" else "Amador Family Correspondence, 1856-1949"
    if value != expected_name:
        return False, None, f"Collection Name does not match expected value for {language}. Expected '{expected_name}' but got '{value}'"
    
    return True, None, "Valid"


def validate_collection_number(value):
    """
    Validates the collection number for both English and Spanish columns.
    The only valid values are 'Ms0004' and 'Ms0071'.

    Parameters:
    - value (str): The collection number to validate.

    Returns:
    - (bool, str, str): Tuple indicating if validation passed, the color for highlighting ('red'), 
      and an error message.
    """
    valid_values = {"Ms0004", "Ms0071"}

    # Clean and normalize the value
    cleaned_value = str(value).strip()

    if cleaned_value in valid_values:
        print(f"Validation successful: Collection number '{cleaned_value}' is valid.")
        return True, "", "Valid collection number"
    else:
        print(f"Failed validation: Collection number '{cleaned_value}' is invalid.")
        return False, "red", f"Invalid collection number: Expected one of {valid_values}, but got '{cleaned_value}'"



def validate_full_folder_or_file_path(value, collection_identifier):
    """
    Validates the FullFolderOrFilePath column for proper structure and naming conventions.

    Parameters:
    - value (str): The full folder or file path to validate.
    - collection_identifier (str): The collection identifier (e.g., 'Ms0004', 'Ms0071').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red' or 'yellow'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating FullFolderOrFilePath value '{value}' with collection identifier '{collection_identifier}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: FullFolderOrFilePath value is empty or missing.")
            return False, "yellow", "FullFolderOrFilePath is empty or missing."

        # Define the standard pattern for the full folder/file path
        standard_pattern = rf"^/Box_\d+/(\d+_\d+)/{collection_identifier}_\d+_\d+_\d+\.pdf$"
        
        # Define the alternate pattern for non-standard numbering
        letter_suffix_pattern = rf"^/Box_\d+/(\d+_\d+)/{collection_identifier}_\d+_\d+_\d+[A-Z]\.pdf$"

        # Check for valid standard format
        if re.match(standard_pattern, value.strip()):
            print("Debug: FullFolderOrFilePath value is valid.")
            return True, "", "Valid FullFolderOrFilePath."

        # Check for valid format but with non-standard numbering
        if re.match(letter_suffix_pattern, value.strip()):
            print("Debug: FullFolderOrFilePath value contains non-standard numbering.")
            return False, "yellow", "Non-standard numbering in file name (e.g., '05A')."

        # If it doesn't match any of the valid patterns
        print("Debug: FullFolderOrFilePath value is invalid.")
        return False, "red", f"Invalid FullFolderOrFilePath format: '{value}'"

    except Exception as e:
        print(f"Error validating FullFolderOrFilePath value '{value}': {e}")
        return False, "red", f"Error validating FullFolderOrFilePath: {e}"




def validate_other_places_mentioned(city, city_info):
    """
    Validates the 'Other Places Mentioned' column by checking city format and existence in the city list.

    Parameters:
    - city (str): City name(s) from the column, separated by '[|]'.
    - city_info (dict): Dictionary containing city data for validation.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red' or 'yellow'), and message.
    """
    try:
        # Debugging log
        print(f"Debug: Starting validation for 'Other Places Mentioned' with value: '{city}'")

        # Clean and validate the city value
        cleaned_city = str(city).strip()
        print(f"Debug: Cleaned city value: '{cleaned_city}'")

        # Split multiple cities using the separator
        cities = cleaned_city.split('[|]')
        invalid_format = []
        not_found = []

        for city in cities:
            city = city.strip()
            # Validate the format (e.g., "CityName (StateAbbr.)")
            if not re.match(r"^[A-Za-z\s]+ \([A-Za-z]+\.\)$", city):
                invalid_format.append(city)
                continue

            # Validate if the city exists in the dataset
            city_key = (city.lower(), 'english')  # Assuming English for simplicity; adjust as needed
            if city_key not in city_info:
                not_found.append(city)

        # Return appropriate validation result
        if invalid_format:
            return False, "red", f"Invalid city format(s): {', '.join(invalid_format)}"
        if not_found:
            return False, "yellow", f"City(s) not found in the approved city list: {', '.join(not_found)}"

        # If everything is valid
        print(f"Debug: Validation passed for cities '{cleaned_city}' in Other Places Mentioned.")
        return True, "", "Valid city names"

    except Exception as e:
        # Debugging log for any unexpected errors
        print(f"Error validating 'Other Places Mentioned' with value '{city}': {e}")
        return False, "red", f"Error validating city: {e}"





def is_valid_date(value):
    if pd.isna(value):
        return True, None, "Valid (empty)"
    try:
        pd.to_datetime(value, format='%Y-%m-%d', errors='raise')
        return True, None, "Valid"
    except (ValueError, TypeError):
        pass
    try:
        pd.to_datetime(value, format='%Y-%m', errors='raise')
        return True, None, "Valid"
    except (ValueError, TypeError):
        return False, None, "Date format is invalid. Expected 'YYYY-MM-DD' or 'YYYY-MM'"


def extract_date_from_title(title):
    """
    Extracts a date from the Title column in the format 'Month Day, Year'
    and converts it to 'YYYY-MM-DD'.
    """
    try:
        # Match patterns like "January 21, 1898"
        match = re.search(r'([A-Za-z]+)\s+(\d{1,2}),\s+(\d{4})', title)
        if match:
            month_str, day, year = match.groups()
            month = datetime.datetime.strptime(month_str, "%B").month  # Convert month name to number
            return f"{year}-{month:02d}-{int(day):02d}"  # Format as YYYY-MM-DD
    except Exception as e:
        print(f"Error extracting date from title '{title}': {e}")
    return None

def validate_date_column(date_value, title_value):
    """
    Validates a date column value against the Title column.
    Ensures:
    1. Correct format: 'YYYY-MM-DD'.
    2. Matches the extracted date from Title column.

    Parameters:
    - date_value (str): The value in the DATE or ES..DATE column.
    - title_value (str): The corresponding Title column value.

    Returns:
    - (bool, str, str): Tuple of validation status, highlight color, and error message.
    """
    # Ensure date_value is in correct format
    try:
        if pd.isna(date_value):
            return False, "yellow", "Date column is empty."

        # Parse date_value to ensure it's in 'YYYY-MM-DD' format
        parsed_date = pd.to_datetime(date_value, format='%Y-%m-%d', errors='coerce')
        if pd.isna(parsed_date):
            return False, "red", "Date format is invalid. Expected 'YYYY-MM-DD'."
    except Exception as e:
        return False, "red", f"Error validating date format: {e}"

    # Extract date from title
    extracted_date = extract_date_from_title(title_value)
    if not extracted_date:
        return False, "yellow", f"Unable to extract date from title: {title_value}"

    # Compare extracted date with date_value
    if extracted_date != date_value.strip():
        return False, "red", f"Date '{date_value}' does not match extracted date '{extracted_date}' from title."

    return True, "", "Date matches and is valid."




def validate_year(year_value, date_value):
    """
    Validates the YEAR and ES..YEAR columns.

    Parameters:
    - year_value (str): The value in the YEAR or ES..YEAR column.
    - date_value (str): The value in the DATE or ES..DATE column.

    Returns:
    - (bool, str, str): Tuple indicating if validation passed, the color for highlighting ('red' or 'yellow'), 
      and an error message.
    """
    try:
        # Ensure year_value is a 4-digit number
        if not year_value.isdigit() or len(year_value) != 4:
            return False, "red", f"Invalid year format: {year_value}"

        # Extract the year from the date_value
        date_year = date_value.split("-")[0] if "-" in date_value else None

        if not date_year:
            return False, "red", f"Invalid date format: {date_value}"

        # Ensure the year matches the year from the date
        if year_value != date_year:
            return False, "yellow", f"Year '{year_value}' does not match date year '{date_year}'"

        return True, "", "Year validation passed"
    except Exception as e:
        return False, "red", f"Error validating year: {str(e)}"




def is_valid_year(value):
    if not isinstance(value, int):
        return False, None, "Year is not an integer"
    if not (1000 <= value <= 9999):
        return False, None, "Year is out of valid range (1000-9999)"
    return True, None, "Valid"



def validate_name_field(value, authorized_names):
    """
    Validates if a name in the FROM or TO fields matches an entry in the authorized names dataset.
    If the name is missing from the dataset entirely, it highlights the cell in yellow.
    If the name exists in the dataset but is not in the correct format, it highlights the cell in red.

    Parameters:
    - value (str): The name to validate.
    - authorized_names (set): Set of authorized names loaded from the dataset.

    Returns:
    - (bool, str, str): Validation status, fill color ('red', 'yellow', or None), and message.
    """
    known_unknown_values = {
        "FROM" : {"Unknown sender"},
        "ES..FROM": {"Remitente desconocido"},
        "TO" : {"Unknown recipient"},
        "ES..TO": {"Destinatario desconocido"}

    }

    # Clean up the input value
    cleaned_value = value.strip()
    
    print(f"Debug: Starting search for '{cleaned_value}' in authorized names...")

    # Check if the value is in the known "unknown" set
    if cleaned_value in known_unknown_values:
        print(f"Debug: '{cleaned_value}' recognized as a known unknown value. Validation passed.")
        return True, None, "Valid (unknown value)"
    
    # Check if cleaned_value matches an authorized name exactly
    if cleaned_value in authorized_names:
        print(f"Debug: '{cleaned_value}' found in authorized names with correct format. Validation passed.")
        return True, None, "Valid"
    
    # Check if a similar name exists but does not match exactly
    matching_names = [name for name in authorized_names if name.lower() == cleaned_value.lower()]
    if matching_names:
        print(f"Debug: '{cleaned_value}' found in authorized names but format is incorrect. Expected '{matching_names[0]}'.")
        return False, "red", f"Format error: Expected '{matching_names[0]}'"
    
    # Name does not exist in the authorized names dataset at all
    print(f"Debug: '{cleaned_value}' not found in authorized names dataset.")
    return False, "yellow", "Name not found in dataset"


def is_valid_subject_lcsh(value, approved_subjects, language="english"):
    if not isinstance(value, str):
        return False, None, "Invalid type: Expected a string"

    terms = [term.strip() for term in value.split("[|]")]
    print(f"Debug: Starting validation for SUBJECT_LCSH terms '{terms}' in language '{language}'...")

    invalid_terms = []
    for term in terms:
        if term in approved_subjects[language]:
            print(f"Debug: Term '{term}' found in vocabulary for '{language}'. Validation passed.")
        else:
            print(f"Debug: Term '{term}' NOT found in vocabulary for '{language}'. Validation failed.")
            invalid_terms.append(term)

    if invalid_terms:
        return False, "yellow", f"Terms not found in vocabulary: {invalid_terms}"
    return True, None, "Valid"

def validate_extent(value, language):
    """
    Validates the EXTENT and ES..EXTENT columns for proper singular/plural usage,
    correct language inside the brackets, and correct format.

    Parameters:
    - value (str): The extent value to validate.
    - language (str): The language of the extent ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red' or 'yellow'), and message.
    """
    try:
        print(f"Debug: Validating EXTENT value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: EXTENT value is empty or missing.")
            return False, "yellow", "Extent value is empty or missing."

        # Define patterns for English and Spanish
        if language == "english":
            pattern = r"^(\d+) (leaf|leaves) \[(\d+) (page|pages)\]$"
        elif language == "spanish":
            pattern = r"^(\d+) (hoja|hojas) \[(\d+) (página|páginas)\]$"
        else:
            return False, "red", f"Unsupported language: {language}"

        match = re.match(pattern, value.strip())
        if not match:
            print(f"Debug: Invalid format detected for EXTENT value '{value}'.")
            return False, "red", f"Invalid format for EXTENT value: '{value}'"

        # Extract leaves/pages and their units
        leaves, leaves_unit, pages, pages_unit = match.groups()
        leaves = int(leaves)
        pages = int(pages)

        # Validate pluralization
        if leaves == 1 and leaves_unit != ("leaf" if language == "english" else "hoja"):
            return False, "red", f"Incorrect singular form: {leaves_unit} for {leaves}."
        elif leaves > 1 and leaves_unit != ("leaves" if language == "english" else "hojas"):
            return False, "red", f"Incorrect plural form: {leaves_unit} for {leaves}."
        if pages == 1 and pages_unit != ("page" if language == "english" else "página"):
            return False, "red", f"Incorrect singular form: {pages_unit} for {pages}."
        elif pages > 1 and pages_unit != ("pages" if language == "english" else "páginas"):
            return False, "red", f"Incorrect plural form: {pages_unit} for {pages}."

        print(f"Debug: Validation passed for EXTENT value '{value}'.")
        return True, "", "Valid EXTENT value"

    except Exception as e:
        print(f"Error validating EXTENT value '{value}': {e}")
        return False, "red", f"Error validating EXTENT value: {e}"


# Allowed values for PHYSICAL_DESCRIPTION and ES..PHYSICAL_DESCRIPTION columns
PHYSICAL_DESCRIPTION_VALUES = {
    "de tinta azul": "blue ink",
    "de tinta morada": "purple ink",
    "de tinta negra": "black ink",
    "de tinta plateada": "silver ink",
    "de tinta verde": "green ink",
    "de tinta dorada": "gold ink",
    "de tinta roja": "red ink",
    "lápiz": "pencil",
    "lápiz azul": "blue pencil",
    "lápiz morado": "purple pencil",
    "crayones": "crayons",
    "escrito a máquina tinta negra": "typewritten black ink",
    "escrito a máquina tinta azul": "typewritten blue ink",
    "escrito a máquina tinta verde": "typewritten green ink",
    "escrito a máquina tinta roja": "typewritten red ink",
    "escrito a máquina tinta morada": "typewritten purple ink",
    "acuarela": "watercolor (paint)",
    "papel cuadriculado": "graph paper",
    "papel en blanco": "blank paper",
    "papel rayado": "ruled paper",
    "papel sin rayas": "unruled paper",
    "papel de color": "colored paper",
    "papelería": "stationery",
    "papelería de luto": "mourning stationery",
    "papel con membrete": "letterheads",
    "papel impreso con tipografía": "letterpress printed paper",
    "cuadernos de tipografía": "letterpress copybooks",
    "sobre": "envelope",
    "tejido": "cloth"
}


def validate_physical_description(value, language):
    """
    Validates the PHYSICAL_DESCRIPTION and ES..PHYSICAL_DESCRIPTION columns.
    Handles multiple terms separated by [|].

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating PHYSICAL_DESCRIPTION value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: PHYSICAL_DESCRIPTION value is empty or missing.")
            return False, "yellow", "Physical description value is empty or missing."

        # Split terms by [|]
        terms = [term.strip() for term in value.split("[|]")]
        print(f"Debug: Parsed terms for validation: {terms}")

        # Validate each term
        if language == "spanish":
            invalid_terms = [term for term in terms if term not in PHYSICAL_DESCRIPTION_VALUES]
            if invalid_terms:
                print(f"Debug: Invalid Spanish terms detected: {invalid_terms}")
                return False, "red", f"Invalid Spanish terms: {invalid_terms}"
        elif language == "english":
            reverse_mapping = {v: k for k, v in PHYSICAL_DESCRIPTION_VALUES.items()}  # Reverse map English to Spanish
            invalid_terms = [term for term in terms if term not in reverse_mapping]
            if invalid_terms:
                print(f"Debug: Invalid English terms detected: {invalid_terms}")
                return False, "red", f"Invalid English terms: {invalid_terms}"
        else:
            return False, "red", f"Unsupported language: {language}"

        # If all checks pass
        print(f"Debug: Validation passed for PHYSICAL_DESCRIPTION value '{value}'.")
        return True, "", "Valid PHYSICAL_DESCRIPTION value"

    except Exception as e:
        print(f"Error validating PHYSICAL_DESCRIPTION value '{value}': {e}")
        return False, "red", f"Error validating PHYSICAL_DESCRIPTION value: {e}"


# Allowed values for DIGITAL_PUBLISHER and ES..DIGITAL_PUBLISHER
DIGITAL_PUBLISHER_ENGLISH = "New Mexico State University Library"
DIGITAL_PUBLISHER_SPANISH = "Biblioteca de la Universidad Estatal de Nuevo México"


def validate_digital_publisher(value, language):
    """
    Validates the DIGITAL_PUBLISHER and ES..DIGITAL_PUBLISHER columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating DIGITAL_PUBLISHER value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: DIGITAL_PUBLISHER value is empty or missing.")
            return False, "yellow", "Digital publisher value is empty or missing."

        value = value.strip()
        if language == "english" and value != DIGITAL_PUBLISHER_ENGLISH:
            print(f"Debug: Invalid English DIGITAL_PUBLISHER value: '{value}'")
            return False, "red", f"Expected '{DIGITAL_PUBLISHER_ENGLISH}', but got '{value}'"
        elif language == "spanish" and value != DIGITAL_PUBLISHER_SPANISH:
            print(f"Debug: Invalid Spanish DIGITAL_PUBLISHER value: '{value}'")
            return False, "red", f"Expected '{DIGITAL_PUBLISHER_SPANISH}', but got '{value}'"

        # If all checks pass
        print(f"Debug: Validation passed for DIGITAL_PUBLISHER value '{value}'.")
        return True, "", "Valid DIGITAL_PUBLISHER value"

    except Exception as e:
        print(f"Error validating DIGITAL_PUBLISHER value '{value}': {e}")
        return False, "red", f"Error validating DIGITAL_PUBLISHER value: {e}"
    


    # Allowed values for SOURCE and ES..SOURCE
SOURCE_ENGLISH = "NMSU Library Archives and Special Collections"
SOURCE_SPANISH = "Archivos y colecciones especiales de la biblioteca de NMSU"

def validate_source(value, language):
    """
    Validates the SOURCE and ES..SOURCE columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating SOURCE value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: SOURCE value is empty or missing.")
            return False, "yellow", "Source value is empty or missing."

        value = value.strip()
        if language == "english" and value != SOURCE_ENGLISH:
            print(f"Debug: Invalid English SOURCE value: '{value}'")
            return False, "red", f"Expected '{SOURCE_ENGLISH}', but got '{value}'"
        elif language == "spanish" and value != SOURCE_SPANISH:
            print(f"Debug: Invalid Spanish SOURCE value: '{value}'")
            return False, "red", f"Expected '{SOURCE_SPANISH}', but got '{value}'"

        # If all checks pass
        print(f"Debug: Validation passed for SOURCE value '{value}'.")
        return True, "", "Valid SOURCE value"

    except Exception as e:
        print(f"Error validating SOURCE value '{value}': {e}")
        return False, "red", f"Error validating SOURCE value: {e}"


# Allowed values for UNIT and ES..UNIT
UNIT_ENGLISH = "Rio Grande Historical Collections"
UNIT_SPANISH = "Colecciones históricas de Río Grande"


def validate_unit(value, language):
    """
    Validates the UNIT and ES..UNIT columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating UNIT value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: UNIT value is empty or missing.")
            return False, "yellow", "Unit value is empty or missing."

        value = value.strip()
        if language == "english" and value != UNIT_ENGLISH:
            print(f"Debug: Invalid English UNIT value: '{value}'")
            return False, "red", f"Expected '{UNIT_ENGLISH}', but got '{value}'"
        elif language == "spanish" and value != UNIT_SPANISH:
            print(f"Debug: Invalid Spanish UNIT value: '{value}'")
            return False, "red", f"Expected '{UNIT_SPANISH}', but got '{value}'"

        # If all checks pass
        print(f"Debug: Validation passed for UNIT value '{value}'.")
        return True, "", "Valid UNIT value"

    except Exception as e:
        print(f"Error validating UNIT value '{value}': {e}")
        return False, "red", f"Error validating UNIT value: {e}"


# Allowed languages for LANGUAGE and ES..LANGUAGE columns
VALID_LANGUAGES_ENGLISH = {"English", "Spanish", "French", "Japanese"}
VALID_LANGUAGES_SPANISH = {"Inglés", "Español", "Francés", "Japonés"}


def validate_language(value, language):
    """
    Validates the LANGUAGE and ES..LANGUAGE columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red' or 'yellow'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating LANGUAGE value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: LANGUAGE value is empty or missing.")
            return False, "yellow", "Language value is empty or missing."

        # Split the languages using the separator "[|]"
        languages = [lang.strip() for lang in value.split("[|]")]

        # Determine the valid set of languages based on column language
        valid_languages = (
            VALID_LANGUAGES_ENGLISH if language == "english" else VALID_LANGUAGES_SPANISH
        )

        # Check each language in the cell
        for lang in languages:
            if lang not in valid_languages:
                print(f"Debug: Invalid language '{lang}' in column '{language}'.")
                return False, "red", f"Invalid language: '{lang}'"

        # If all languages are valid
        print(f"Debug: All languages '{languages}' are valid in column '{language}'.")
        return True, "", "Valid LANGUAGE value"

    except Exception as e:
        print(f"Error validating LANGUAGE value '{value}': {e}")
        return False, "red", f"Error validating LANGUAGE value: {e}"


# Allowed values for FORMAT and ES..FORMAT
FORMAT_ENGLISH = "application/pdf"
FORMAT_SPANISH = "la aplicación/pdf"

def validate_format(value, language):
    """
    Validates the FORMAT and ES..FORMAT columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating FORMAT value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: FORMAT value is empty or missing.")
            return False, "yellow", "Format value is empty or missing."

        value = value.strip()
        if language == "english" and value != FORMAT_ENGLISH:
            print(f"Debug: Invalid English FORMAT value: '{value}'")
            return False, "red", f"Expected '{FORMAT_ENGLISH}', but got '{value}'"
        elif language == "spanish" and value != FORMAT_SPANISH:
            print(f"Debug: Invalid Spanish FORMAT value: '{value}'")
            return False, "red", f"Expected '{FORMAT_SPANISH}', but got '{value}'"

        # If all checks pass
        print(f"Debug: Validation passed for FORMAT value '{value}'.")
        return True, "", "Valid FORMAT value"

    except Exception as e:
        print(f"Error validating FORMAT value '{value}': {e}")
        return False, "red", f"Error validating FORMAT value: {e}"


# Allowed values for TYPE and ES..TYPE
TYPE_ENGLISH = "Text"
TYPE_SPANISH = "Texto"

def validate_type(value, language):
    """
    Validates the TYPE and ES..TYPE columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating TYPE value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: TYPE value is empty or missing.")
            return False, "yellow", "Type value is empty or missing."

        value = value.strip()
        if language == "english" and value != TYPE_ENGLISH:
            print(f"Debug: Invalid English TYPE value: '{value}'")
            return False, "red", f"Expected '{TYPE_ENGLISH}', but got '{value}'"
        elif language == "spanish" and value != TYPE_SPANISH:
            print(f"Debug: Invalid Spanish TYPE value: '{value}'")
            return False, "red", f"Expected '{TYPE_SPANISH}', but got '{value}'"

        # If all checks pass
        print(f"Debug: Validation passed for TYPE value '{value}'.")
        return True, "", "Valid TYPE value"

    except Exception as e:
        print(f"Error validating TYPE value '{value}': {e}")
        return False, "red", f"Error validating TYPE value: {e}"



# Allowed values for MEDIUM_AAT and ES..MEDIUM_AAT
MEDIUM_ENGLISH = {
    "correspondence artifacts",
    "personal correspondence",
    "commercial correspondence",
    "legal correspondence"
}
MEDIUM_SPANISH = {
    "artefactos de correspondencia",
    "correspondencia personal",
    "correspondencia comercial",
    "correspondencia legal"
}


def validate_medium(value, language):
    """
    Validates the MEDIUM_AAT and ES..MEDIUM_AAT columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating MEDIUM value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: MEDIUM value is empty or missing.")
            return False, "yellow", "Medium value is empty or missing."

        value = value.strip()
        valid_values = MEDIUM_ENGLISH if language == "english" else MEDIUM_SPANISH

        if value not in valid_values:
            print(f"Debug: Invalid {language.upper()} MEDIUM value: '{value}'")
            return False, "red", f"Invalid medium: '{value}'. Expected one of {valid_values}"

        # If all checks pass
        print(f"Debug: Validation passed for MEDIUM value '{value}'.")
        return True, "", "Valid MEDIUM value"

    except Exception as e:
        print(f"Error validating MEDIUM value '{value}': {e}")
        return False, "red", f"Error validating MEDIUM value: {e}"
    
# Allowed values for GENRE_AAT and ES..GENRE_AAT
GENRE_VALUES = {
    "advertisements": "anuncios publicitarios",
    "announcements": "anuncios",
    "anniversary announcements": "anuncios de aniversario",
    "wedding announcements": "anuncios de boda",
    "graduation announcements": "anuncios de graduación",
    "birth announcements": "anuncios de nacimiento",
    "business announcements": "anuncios comerciales",
    "award announcements": "anuncios de premios",
    "funeral announcements": "anuncios funerarios",
    "business letters": "cartas comerciales",
    "invitations": "invitaciones",
    "manuscripts (documents)": "manuscritos",
    "typescripts": "manuscritos dactilografiados",
    "picture postcards": "tarjetas postales ilustradas",
    "postcards": "tarjetas postales",
    "telegrams": "telegramas",
    "programs (documents)": "programas",
    "concert programs": "programas de conciertos",
    "dance cards": "tarjetas de baile",
    "birthday cards": "tarjetas de cumpleaños",
    "holiday cards": "tarjetas de festividades",
    "memorial cards": "tarjetas de memorial",
    "devotional cards": "tarjetas devocionales",
    "get well cards": "tarjetas de mejoría",
    "greeting cards": "tarjetas de felicitación",
    "clippings (information artifacts)": "recortes",
    "sympathy cards": "tarjetas de pésame",
    "brochures": "folletos (publicidad)",
    "business cards": "tarjetas comerciales",
    "notes (documents)": "notas",
    "social cards": "tarjetas sociales",
    "tickets": "billetes",
    "prescriptions": "prescripciones",
    "receipts (financial records)": "recibo (carta de pago)"
}


def validate_genre(value, language):
    """
    Validates the GENRE_AAT and ES..GENRE_AAT columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating GENRE value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: GENRE value is empty or missing.")
            return False, "yellow", "Genre value is empty or missing."

        value = value.strip()
        valid_values = GENRE_VALUES.keys() if language == "english" else GENRE_VALUES.values()

        if value not in valid_values:
            print(f"Debug: Invalid {language.upper()} GENRE value: '{value}'")
            return False, "red", f"Invalid genre: '{value}'. Expected one of {list(valid_values)}"

        # If all checks pass
        print(f"Debug: Validation passed for GENRE value '{value}'.")
        return True, "", "Valid GENRE value"

    except Exception as e:
        print(f"Error validating GENRE value '{value}': {e}")
        return False, "red", f"Error validating GENRE value: {e}"
    

# Allowed values for ACCESS_RIGHTS and ES..ACCESS_RIGHTS
ACCESS_RIGHTS_ENGLISH = "Open for re-use"
ACCESS_RIGHTS_SPANISH = "Abierto para la reutilización"

def validate_access_rights(value, language):
    """
    Validates the ACCESS_RIGHTS and ES..ACCESS_RIGHTS columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating ACCESS_RIGHTS value '{value}' in language '{language}'")

        if pd.isna(value) or str(value).strip() == "":
            print("Debug: ACCESS_RIGHTS value is empty or missing.")
            return False, "yellow", "Access rights value is empty or missing."

        value = value.strip()
        expected_value = ACCESS_RIGHTS_ENGLISH if language == "english" else ACCESS_RIGHTS_SPANISH

        if value != expected_value:
            print(f"Debug: Invalid {language.upper()} ACCESS_RIGHTS value: '{value}'")
            return False, "red", f"Invalid access rights: '{value}'. Expected '{expected_value}'"

        # If all checks pass
        print(f"Debug: Validation passed for ACCESS_RIGHTS value '{value}'.")
        return True, "", "Valid ACCESS_RIGHTS value"

    except Exception as e:
        print(f"Error validating ACCESS_RIGHTS value '{value}': {e}")
        return False, "red", f"Error validating ACCESS_RIGHTS value: {e}"


def validate_metadata_cataloger(value):
    """
    Validates the METADATA_CATALOGER and ES..METADATA_CATALOGER columns.

    Parameters:
    - value (str): The value to validate.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating METADATA_CATALOGER value '{value}'")

        # Ensure the value is not empty or NaN
        if pd.isna(value) or str(value).strip() == "":
            return False, "yellow", "Metadata cataloger value is empty or missing."

        value = value.strip()

        # Regular expression to match 'LastName, FirstName' format
        if not re.match(r"^[A-Za-z]+, [A-Za-z]+$", value):
            return False, "red", f"Invalid format: '{value}'. Expected 'LastName, FirstName'."

        # Validation passed
        print(f"Debug: Validation passed for METADATA_CATALOGER value '{value}'.")
        return True, "", "Valid METADATA_CATALOGER value"

    except Exception as e:
        print(f"Error validating METADATA_CATALOGER value '{value}': {e}")
        return False, "red", f"Error validating METADATA_CATALOGER value: {e}"


# Allowed values for OA_DESCRIPTION and ES..OA_DESCRIPTION
OA_DESCRIPTION_ENGLISH = "This collection is available in both, English and Spanish"
OA_DESCRIPTION_SPANISH = "Esta colección está disponible en inglés y español"


def validate_oa_description(value, language):
    """
    Validates the OA_DESCRIPTION and ES..OA_DESCRIPTION columns.

    Parameters:
    - value (str): The value to validate.
    - language (str): The language of the column ('english' or 'spanish').

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating OA_DESCRIPTION value '{value}' in language '{language}'")

        # Ensure the value is not empty or NaN
        if pd.isna(value) or str(value).strip() == "":
            return False, "yellow", "OA_DESCRIPTION value is empty or missing."

        value = value.strip()
        expected_value = OA_DESCRIPTION_ENGLISH if language == "english" else OA_DESCRIPTION_SPANISH

        if value != expected_value:
            print(f"Debug: Invalid {language.upper()} OA_DESCRIPTION value: '{value}'")
            return False, "red", f"Invalid OA_DESCRIPTION: '{value}'. Expected '{expected_value}'"

        # Validation passed
        print(f"Debug: Validation passed for OA_DESCRIPTION value '{value}'.")
        return True, "", "Valid OA_DESCRIPTION value"

    except Exception as e:
        print(f"Error validating OA_DESCRIPTION value '{value}': {e}")
        return False, "red", f"Error validating OA_DESCRIPTION value: {e}"
    
# Allowed value for OA_COLLECTION
OA_COLLECTION_VALID_VALUE = "10317"

def validate_oa_collection(value):
    """
    Validates the OA_COLLECTION column.

    Parameters:
    - value (str): The value to validate.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and a validation message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating OA_COLLECTION value '{value}'")

        # Ensure the value is not empty or NaN
        if pd.isna(value) or str(value).strip() == "":
            return False, "yellow", "OA_COLLECTION value is empty or missing."

        value = str(value).strip()

        if value != OA_COLLECTION_VALID_VALUE:
            print(f"Debug: Invalid OA_COLLECTION value: '{value}'. Expected '{OA_COLLECTION_VALID_VALUE}'.")
            return False, "red", f"Invalid OA_COLLECTION: '{value}'. Expected '{OA_COLLECTION_VALID_VALUE}'."

        # Validation passed
        print(f"Debug: Validation passed for OA_COLLECTION value '{value}'.")
        return True, "", "Valid OA_COLLECTION value"

    except Exception as e:
        print(f"Error validating OA_COLLECTION value '{value}': {e}")
        return False, "red", f"Error validating OA_COLLECTION value: {e}"

# Allowed value for OA_PROFILE
OA_PROFILE_VALID_VALUE = "Documents"

def validate_oa_profile(value):
    """
    Validates the OA_PROFILE column.

    Parameters:
    - value (str): The value to validate.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and a validation message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating OA_PROFILE value '{value}'")

        # Ensure the value is not empty or NaN
        if pd.isna(value) or str(value).strip() == "":
            return False, "yellow", "OA_PROFILE value is empty or missing."

        value = str(value).strip()

        if value != OA_PROFILE_VALID_VALUE:
            print(f"Debug: Invalid OA_PROFILE value: '{value}'. Expected '{OA_PROFILE_VALID_VALUE}'.")
            return False, "red", f"Invalid OA_PROFILE: '{value}'. Expected '{OA_PROFILE_VALID_VALUE}'."

        # Validation passed
        print(f"Debug: Validation passed for OA_PROFILE value '{value}'.")
        return True, "", "Valid OA_PROFILE value"

    except Exception as e:
        print(f"Error validating OA_PROFILE value '{value}': {e}")
        return False, "red", f"Error validating OA_PROFILE value: {e}"


# Allowed value for OA_STATUS
OA_STATUS_VALID_VALUE = "PUBLISH"

def validate_oa_status(value):
    """
    Validates the OA_STATUS column.

    Parameters:
    - value (str): The value to validate.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and a validation message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating OA_STATUS value '{value}'")

        # Ensure the value is not empty or NaN
        if pd.isna(value) or str(value).strip() == "":
            return False, "yellow", "OA_STATUS value is empty or missing."

        value = str(value).strip()

        if value != OA_STATUS_VALID_VALUE:
            print(f"Debug: Invalid OA_STATUS value: '{value}'. Expected '{OA_STATUS_VALID_VALUE}'.")
            return False, "red", f"Invalid OA_STATUS: '{value}'. Expected '{OA_STATUS_VALID_VALUE}'."

        # Validation passed
        print(f"Debug: Validation passed for OA_STATUS value '{value}'.")
        return True, "", "Valid OA_STATUS value"

    except Exception as e:
        print(f"Error validating OA_STATUS value '{value}': {e}")
        return False, "red", f"Error validating OA_STATUS value: {e}"

# Allowed value for OA_OBJECT_TYPE
OA_OBJECT_TYPE_VALID_VALUE = "RECORD"

def validate_oa_object_type(value):
    """
    Validates the OA_OBJECT_TYPE column.

    Parameters:
    - value (str): The value to validate.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and a validation message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating OA_OBJECT_TYPE value '{value}'")

        # Ensure the value is not empty or NaN
        if pd.isna(value) or str(value).strip() == "":
            return False, "yellow", "OA_OBJECT_TYPE value is empty or missing."

        value = str(value).strip()

        if value != OA_OBJECT_TYPE_VALID_VALUE:
            print(f"Debug: Invalid OA_OBJECT_TYPE value: '{value}'. Expected '{OA_OBJECT_TYPE_VALID_VALUE}'.")
            return False, "red", f"Invalid OA_OBJECT_TYPE: '{value}'. Expected '{OA_OBJECT_TYPE_VALID_VALUE}'."

        # Validation passed
        print(f"Debug: Validation passed for OA_OBJECT_TYPE value '{value}'.")
        return True, "", "Valid OA_OBJECT_TYPE value"

    except Exception as e:
        print(f"Error validating OA_OBJECT_TYPE value '{value}': {e}")
        return False, "red", f"Error validating OA_OBJECT_TYPE value: {e}"
    
# Allowed value for OA_METADATA_SCHEMA
OA_METADATA_SCHEMA_VALID_VALUE = "4"

def validate_oa_metadata_schema(value):
    """
    Validates the OA_METADATA_SCHEMA column.

    Parameters:
    - value (str): The value to validate.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and a validation message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating OA_METADATA_SCHEMA value '{value}'")

        # Ensure the value is not empty or NaN
        if pd.isna(value) or str(value).strip() == "":
            return False, "yellow", "OA_METADATA_SCHEMA value is empty or missing."

        value = str(value).strip()

        if value != OA_METADATA_SCHEMA_VALID_VALUE:
            print(f"Debug: Invalid OA_METADATA_SCHEMA value: '{value}'. Expected '{OA_METADATA_SCHEMA_VALID_VALUE}'.")
            return False, "red", f"Invalid OA_METADATA_SCHEMA: '{value}'. Expected '{OA_METADATA_SCHEMA_VALID_VALUE}'."

        # Validation passed
        print(f"Debug: Validation passed for OA_METADATA_SCHEMA value '{value}'.")
        return True, "", "Valid OA_METADATA_SCHEMA value"

    except Exception as e:
        print(f"Error validating OA_METADATA_SCHEMA value '{value}': {e}")
        return False, "red", f"Error validating OA_METADATA_SCHEMA value: {e}"
    

# Allowed value for OA_FEATURED
OA_FEATURED_VALID_VALUE = "0"

def validate_oa_featured(value):
    """
    Validates the OA_FEATURED column.

    Parameters:
    - value (str): The value to validate.

    Returns:
    - (bool, str, str): Validation status, highlight color ('red'), and a validation message.
    """
    try:
        # Debug log for the function call
        print(f"Debug: Validating OA_FEATURED value '{value}'")

        # Ensure the value is not empty or NaN
        if pd.isna(value) or str(value).strip() == "":
            return False, "yellow", "OA_FEATURED value is empty or missing."

        value = str(value).strip()

        if value != OA_FEATURED_VALID_VALUE:
            print(f"Debug: Invalid OA_FEATURED value: '{value}'. Expected '{OA_FEATURED_VALID_VALUE}'.")
            return False, "red", f"Invalid OA_FEATURED: '{value}'. Expected '{OA_FEATURED_VALID_VALUE}'."

        # Validation passed
        print(f"Debug: Validation passed for OA_FEATURED value '{value}'.")
        return True, "", "Valid OA_FEATURED value"

    except Exception as e:
        print(f"Error validating OA_FEATURED value '{value}': {e}")
        return False, "red", f"Error validating OA_FEATURED value: {e}"



def load_authorized_names(names_dataset_path):
    names_data = pd.read_excel(names_dataset_path, usecols=[0], header=None)
    authorized_names = set(names_data[0].dropna().str.strip())
    print("Debug: Authorized names loaded from dataset:", authorized_names)
    return authorized_names

# Column validation rules
column_validation_rules = {
    "DIGITAL_IDENTIFIER": validate_digital_identifier,
    "ES..DIGITAL_IDENTIFIER": validate_digital_identifier,
    "BOX_FOLDER": is_valid_box_folder,
    "ES..BOX_FOLDER": is_valid_box_folder,
    "COLLECTION_NAME": lambda x: is_valid_collection_name(x, language="English"),
    "ES..COLLECTION_NAME": lambda x: is_valid_collection_name(x, language="Spanish"),
    "DATE": is_valid_date,
    "ES..DATE": is_valid_date,
    "YEAR": is_valid_year,
    "ES..YEAR": is_valid_year,
    "SUBJECT_LCSH": lambda x: is_valid_subject_lcsh(x, approved_subjects, language="english"),
    "ES..SUBJECT_LCSH" : lambda x: is_valid_subject_lcsh(x, approved_subjects, language="spanish"),
    "FROM": lambda x: validate_name_field(x, authorized_names),
    "ES..FROM": lambda x: validate_name_field(x, authorized_names),
    "TO": lambda x: validate_name_field(x, authorized_names),
    "ES..TO": lambda x: validate_name_field(x, authorized_names),
    # Add validation rules to the main column validation structure
    "SERIES" : validate_series,
    "ES..SERIES" : validate_series
}

location_validation_rules = {
    "SENDERS_CITY": lambda row: is_valid_city_related(row, 'SENDERS_CITY', 'SENDERS_COUNTRY', 'SENDERS_STATE', 'GEOLOC_SCITY', 'english'),
    "ES..SENDERS_CITY": lambda row: is_valid_city_related(row, 'ES..SENDERS_CITY', 'ES..SENDERS_COUNTRY', 'ES..SENDERS_STATE', 'ES..GEOLOC_SCITY', 'spanish'),
    "ADDRESSEES_CITY": lambda row: is_valid_city_related(row, 'ADDRESSEES_CITY', 'ADDRESSEES_COUNTRY', 'ADDRESSEES_STATE', 'GEOLOC_SCITY', 'english'),
    "ES..ADDRESSEES_CITY": lambda row: is_valid_city_related(row, 'ES..ADDRESSEES_CITY', 'ES..ADDRESSEES_COUNTRY', 'ES..ADDRESSEES_STATE', 'ES..GEOLOC_SCITY', 'spanish')
}



approved_subjects = load_approved_subjects("SUBJECT_LCSH.xlsx")
city_info = load_city_data("Maybeee.xlsx")
authorized_names = load_authorized_names("CVPeople.xlsx")
print("Loaded authorized names:", authorized_names)

# The `verify_file` function and main script setup remain the same, using `column_validation_rules`.

def verify_file(input_file, output_file):
   # Load the workbook
    wb = load_workbook(input_file)

    # Get the first sheet name dynamically
    sheet_name = wb.sheetnames[0]

    # Load the first sheet
    ws = wb[sheet_name]
    df = pd.read_excel(input_file, sheet_name=sheet_name)

    # Initialize previous identifiers separately for each column
    previous_identifier_digital = None  # For DIGITAL_IDENTIFIER column
    previous_identifier_es_digital = None  # For ES..DIGITAL_IDENTIFIER column

    for idx, row in df.iterrows():
        # Validate non-location columns
        for col_name, validation_func in column_validation_rules.items():
            if col_name in df.columns:
                value = row[col_name]
                try:
                    # Special handling for DIGITAL_IDENTIFIER to track sequence
                    if col_name == "DIGITAL_IDENTIFIER":
                        is_valid, result, message = validate_digital_identifier(value, previous_identifier_digital)
                        if is_valid:
                            previous_identifier_digital = result  # Update tracking for next row
                            print(f"Validation successful: {col_name} at row {idx + 2}")
                        else:
                            # Apply the correct color highlight based on the validation result
                            col_idx = df.columns.get_loc(col_name) + 1
                            ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if result == "red" else highlight_fill_yellow
                            print(f"Failed validation: {col_name} at row {idx + 2} - Reason: {message}")
                    
                    elif col_name == "ES..DIGITAL_IDENTIFIER":
                        is_valid, result, message = validate_digital_identifier(value, previous_identifier_es_digital)
                        if is_valid:
                            previous_identifier_es_digital = result  # Update tracking for next row
                            print(f"Validation successful: {col_name} at row {idx + 2}")
                        else:
                            col_idx = df.columns.get_loc(col_name) + 1
                            ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if result == "red" else highlight_fill_yellow
                            print(f"Failed validation: {col_name} at row {idx + 2} - Reason: {message}")

                    else:
                        # General validation for non-DIGITAL_IDENTIFIER columns
                        is_valid, color, message = validation_func(value)
                        if is_valid:
                            print(f"Validation successful: {col_name} at row {idx + 2}")
                        else:
                            col_idx = df.columns.get_loc(col_name) + 1
                            ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                            print(f"Failed validation: {col_name} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col_name} at row {idx + 2}: {e}")
        
        # Validate SERIES and ES..SERIES columns
        if "SERIES" in df.columns:
            value = row["SERIES"]
            try:
                is_valid, color, message = validate_series(value, series_values)
                if is_valid:
                    print(f"Validation successful: SERIES at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("SERIES") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: SERIES at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating SERIES at row {idx + 2}: {e}")

        if "ES..SERIES" in df.columns:
            value = row["ES..SERIES"]
            try:
                is_valid, color, message = validate_series(value, series_values)
                if is_valid:
                    print(f"Validation successful: ES..SERIES at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("ES..SERIES") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: ES..SERIES at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating ES..SERIES at row {idx + 2}: {e}")

        # Validate COLLECTION_NUMBER and ES..COLLECTION_NUMBER using validate_collection_number
        if "COLLECTION_NUMBER" in df.columns:
            value = row["COLLECTION_NUMBER"]
            try:
                is_valid, color, message = validate_collection_number(value)
                if is_valid:
                    print(f"Validation successful: COLLECTION_NUMBER at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("COLLECTION_NUMBER") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: COLLECTION_NUMBER at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating COLLECTION_NUMBER at row {idx + 2}: {e}")

        if "ES..COLLECTION_NUMBER" in df.columns:
            value = row["ES..COLLECTION_NUMBER"]
            try:
                is_valid, color, message = validate_collection_number(value)
                if is_valid:
                    print(f"Validation successful: ES..COLLECTION_NUMBER at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("ES..COLLECTION_NUMBER") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: ES..COLLECTION_NUMBER at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating ES..COLLECTION_NUMBER at row {idx + 2}: {e}")

        # Validate COLLECTION_NAME and ES..COLLECTION_NAME
        if "COLLECTION_NAME" in df.columns:
            value = row["COLLECTION_NAME"]
            try:
                is_valid, color, message = validate_collection_name(value, "English")
                if is_valid:
                    print(f"Validation successful: COLLECTION_NAME at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("COLLECTION_NAME") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: COLLECTION_NAME at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating COLLECTION_NAME at row {idx + 2}: {e}")
        
        if "ES..COLLECTION_NAME" in df.columns:
            value = row["ES..COLLECTION_NAME"]
            try:
                is_valid, color, message = validate_collection_name(value, "Spanish")
                if is_valid:
                    print(f"Validation successful: ES..COLLECTION_NAME at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("ES..COLLECTION_NAME") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: ES..COLLECTION_NAME at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating ES..COLLECTION_NAME at row {idx + 2}: {e}")

        # Validate DATE and ES..DATE columns
        if "DATE" in df.columns:
            date_value = row["DATE"]
            title_value = row.get("TITLE", "")  # Ensure Title column exists
            try:
                is_valid, color, message = validate_date_column(date_value, title_value)
                if is_valid:
                    print(f"Validation successful: DATE at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("DATE") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: DATE at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating DATE at row {idx + 2}: {e}")

        if "ES..DATE" in df.columns:
            date_value = row["ES..DATE"]
            title_value = row.get("TITLE", "")  # Ensure Title column exists
            try:
                is_valid, color, message = validate_date_column(date_value, title_value)
                if is_valid:
                    print(f"Validation successful: ES..DATE at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("ES..DATE") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: ES..DATE at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating ES..DATE at row {idx + 2}: {e}")

        # Validate YEAR column
        if "YEAR" in df.columns and "DATE" in df.columns:
            year_value = row["YEAR"]
            date_value = row["DATE"]
            try:
                is_valid, color, message = validate_year(str(year_value), str(date_value))
                if is_valid:
                    print(f"Validation successful: YEAR at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("YEAR") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: YEAR at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating YEAR at row {idx + 2}: {e}")

        # Validate ES..YEAR column
        if "ES..YEAR" in df.columns and "ES..DATE" in df.columns:
            year_value = row["ES..YEAR"]
            date_value = row["ES..DATE"]
            try:
                is_valid, color, message = validate_year(str(year_value), str(date_value))
                if is_valid:
                    print(f"Validation successful: ES..YEAR at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("ES..YEAR") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: ES..YEAR at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating ES..YEAR at row {idx + 2}: {e}")

        # Validate RELATIONSHIP columns
        if "RELATIONSHIP1" in df.columns and "RELATIONSHIP2" in df.columns:
            try:
                is_valid, color, message = validate_relationships(
                    row["RELATIONSHIP1"], row["RELATIONSHIP2"],
                    relationship_mapping_english, "English", "RELATIONSHIP1", "RELATIONSHIP2"
                )
                if is_valid:
                    print(f"Validation successful: RELATIONSHIP1 and RELATIONSHIP2 at row {idx + 2}")
                else:
                    col_idx_rel1 = df.columns.get_loc("RELATIONSHIP1") + 1
                    col_idx_rel2 = df.columns.get_loc("RELATIONSHIP2") + 1
                    ws.cell(row=idx + 2, column=col_idx_rel1).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    ws.cell(row=idx + 2, column=col_idx_rel2).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: RELATIONSHIP1 and RELATIONSHIP2 at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating RELATIONSHIP1 and RELATIONSHIP2 at row {idx + 2}: {e}")

        if "ES..RELATIONSHIP1" in df.columns and "ES..RELATIONSHIP2" in df.columns:
            try:
                is_valid, color, message = validate_relationships(
                    row["ES..RELATIONSHIP1"], row["ES..RELATIONSHIP2"],
                    relationship_mapping, "Spanish", "ES..RELATIONSHIP1", "ES..RELATIONSHIP2"
                )
                if is_valid:
                    print(f"Validation successful: ES..RELATIONSHIP1 and ES..RELATIONSHIP2 at row {idx + 2}")
                else:
                    col_idx_rel1 = df.columns.get_loc("ES..RELATIONSHIP1") + 1
                    col_idx_rel2 = df.columns.get_loc("ES..RELATIONSHIP2") + 1
                    ws.cell(row=idx + 2, column=col_idx_rel1).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    ws.cell(row=idx + 2, column=col_idx_rel2).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                    print(f"Failed validation: ES..RELATIONSHIP1 and ES..RELATIONSHIP2 at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating ES..RELATIONSHIP1 and ES..RELATIONSHIP2 at row {idx + 2}: {e}")

        if "FullFolderOrFilePath" in df.columns:
            for idx, row in df.iterrows():
                full_folder_value = row["FullFolderOrFilePath"]
                digital_identifier = row.get("DIGITAL_IDENTIFIER", "")

                # Extract collection identifier (e.g., 'Ms0004' or 'Ms0071') from DIGITAL_IDENTIFIER
                collection_identifier = digital_identifier.split("_")[0] if "_" in digital_identifier else digital_identifier

                try:
                    is_valid, color, message = validate_full_folder_or_file_path(full_folder_value, collection_identifier)
                    if is_valid:
                        print(f"Validation successful: FullFolderOrFilePath at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc("FullFolderOrFilePath") + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                        print(f"Failed validation: FullFolderOrFilePath at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating FullFolderOrFilePath at row {idx + 2}: {e}")



      # Validate 'OTHER_PLACES_MENTIONED' and 'ES..OTHER_PLACES_MENTIONED' columns
        for col in ["OTHER_PLACES_MENTIONED", "ES..OTHER_PLACES_MENTIONED"]:
            if col in df.columns:
                # Debugging log to confirm the column exists and the verification is starting
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    # Use the validate function for city names
                    is_valid, color, message = validate_other_places_mentioned(value, city_info)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                # Debugging log to confirm the column is not found
                print(f"Debug: Column '{col}' not found in dataset.")


        # Validate 'EXTENT' and 'ES..EXTENT' columns
        for col, lang in [("EXTENT", "english"), ("ES..EXTENT", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_extent(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")



        # Validate 'PHYSICAL_DESCRIPTION' and 'ES..PHYSICAL_DESCRIPTION' columns
        for col, lang in [("PHYSICAL_DESCRIPTION", "english"), ("ES..PHYSICAL_DESCRIPTION", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_physical_description(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'DIGITAL_PUBLISHER' and 'ES..DIGITAL_PUBLISHER' columns
        for col, lang in [("DIGITAL_PUBLISHER", "english"), ("ES..DIGITAL_PUBLISHER", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_digital_publisher(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'SOURCE' and 'ES..SOURCE' columns
        for col, lang in [("SOURCE", "english"), ("ES..SOURCE", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_source(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'UNIT' and 'ES..UNIT' columns
        for col, lang in [("UNIT", "english"), ("ES..UNIT", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_unit(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")


        # Validate 'LANGUAGE' and 'ES..LANGUAGE' columns
        for col, lang in [("LANGUAGE", "english"), ("ES..LANGUAGE", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_language(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'FORMAT' and 'ES..FORMAT' columns
        for col, lang in [("FORMAT", "english"), ("ES..FORMAT", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_format(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'TYPE' and 'ES..TYPE' columns
        for col, lang in [("TYPE", "english"), ("ES..TYPE", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_type(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'MEDIUM_AAT' and 'ES..MEDIUM_AAT' columns
        for col, lang in [("MEDIUM_AAT", "english"), ("ES..MEDIUM_AAT", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_medium(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'GENRE_AAT' and 'ES..GENRE_AAT' columns
        for col, lang in [("GENRE_AAT", "english"), ("ES..GENRE_AAT", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_genre(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'ACCESS_RIGHTS' and 'ES..ACCESS_RIGHTS' columns
        for col, lang in [("ACCESS_RIGHTS", "english"), ("ES..ACCESS_RIGHTS", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_access_rights(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'METADATA_CATALOGER' and 'ES..METADATA_CATALOGER' columns
        for col in ["METADATA_CATALOGER", "ES..METADATA_CATALOGER"]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_metadata_cataloger(value)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'OA_DESCRIPTION' and 'ES..OA_DESCRIPTION' columns
        for col, lang in [("OA_DESCRIPTION", "english"), ("ES..OA_DESCRIPTION", "spanish")]:
            if col in df.columns:
                print(f"Debug: Starting verification for '{col}' column at row {idx + 2}")

                value = row[col]
                # Skip validation if the cell is empty
                if pd.isna(value) or str(value).strip() == "":
                    print(f"Debug: '{col}' at row {idx + 2} is empty. Skipping validation.")
                    continue

                try:
                    is_valid, color, message = validate_oa_description(value, lang)
                    if is_valid:
                        print(f"Validation successful: {col} at row {idx + 2}")
                    else:
                        col_idx = df.columns.get_loc(col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                        print(f"Failed validation: {col} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error validating {col} at row {idx + 2}: {e}")
            else:
                print(f"Debug: Column '{col}' not found in dataset.")

        # Validate 'OA_COLLECTION' column
        if "OA_COLLECTION" in df.columns:
            print(f"Debug: Starting verification for 'OA_COLLECTION' column at row {idx + 2}")

            value = row["OA_COLLECTION"]
            try:
                is_valid, color, message = validate_oa_collection(value)
                if is_valid:
                    print(f"Validation successful: OA_COLLECTION at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("OA_COLLECTION") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                    print(f"Failed validation: OA_COLLECTION at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating OA_COLLECTION at row {idx + 2}: {e}")
        else:
            print("Debug: Column 'OA_COLLECTION' not found in dataset.")

        # Validate 'OA_PROFILE' column
        if "OA_PROFILE" in df.columns:
            print(f"Debug: Starting verification for 'OA_PROFILE' column at row {idx + 2}")

            value = row["OA_PROFILE"]
            try:
                is_valid, color, message = validate_oa_profile(value)
                if is_valid:
                    print(f"Validation successful: OA_PROFILE at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("OA_PROFILE") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                    print(f"Failed validation: OA_PROFILE at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating OA_PROFILE at row {idx + 2}: {e}")
        else:
            print("Debug: Column 'OA_PROFILE' not found in dataset.")

        # Validate 'OA_STATUS' column
        if "OA_STATUS" in df.columns:
            print(f"Debug: Starting verification for 'OA_STATUS' column at row {idx + 2}")

            value = row["OA_STATUS"]
            try:
                is_valid, color, message = validate_oa_status(value)
                if is_valid:
                    print(f"Validation successful: OA_STATUS at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("OA_STATUS") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                    print(f"Failed validation: OA_STATUS at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating OA_STATUS at row {idx + 2}: {e}")
        else:
            print("Debug: Column 'OA_STATUS' not found in dataset.")

        # Validate 'OA_OBJECT_TYPE' column
        if "OA_OBJECT_TYPE" in df.columns:
            print(f"Debug: Starting verification for 'OA_OBJECT_TYPE' column at row {idx + 2}")

            value = row["OA_OBJECT_TYPE"]
            try:
                is_valid, color, message = validate_oa_object_type(value)
                if is_valid:
                    print(f"Validation successful: OA_OBJECT_TYPE at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("OA_OBJECT_TYPE") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                    print(f"Failed validation: OA_OBJECT_TYPE at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating OA_OBJECT_TYPE at row {idx + 2}: {e}")
        else:
            print("Debug: Column 'OA_OBJECT_TYPE' not found in dataset.")

        # Validate 'OA_METADATA_SCHEMA' column
        if "OA_METADATA_SCHEMA" in df.columns:
            print(f"Debug: Starting verification for 'OA_METADATA_SCHEMA' column at row {idx + 2}")

            value = row["OA_METADATA_SCHEMA"]
            try:
                is_valid, color, message = validate_oa_metadata_schema(value)
                if is_valid:
                    print(f"Validation successful: OA_METADATA_SCHEMA at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("OA_METADATA_SCHEMA") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                    print(f"Failed validation: OA_METADATA_SCHEMA at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating OA_METADATA_SCHEMA at row {idx + 2}: {e}")
        else:
            print("Debug: Column 'OA_METADATA_SCHEMA' not found in dataset.")

        # Validate 'OA_FEATURED' column
        if "OA_FEATURED" in df.columns:
            print(f"Debug: Starting verification for 'OA_FEATURED' column at row {idx + 2}")

            value = row["OA_FEATURED"]
            try:
                is_valid, color, message = validate_oa_featured(value)
                if is_valid:
                    print(f"Validation successful: OA_FEATURED at row {idx + 2}")
                else:
                    col_idx = df.columns.get_loc("OA_FEATURED") + 1
                    ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red
                    print(f"Failed validation: OA_FEATURED at row {idx + 2} - Reason: {message}")
            except Exception as e:
                print(f"Error validating OA_FEATURED at row {idx + 2}: {e}")
        else:
            print("Debug: Column 'OA_FEATURED' not found in dataset.")





        # Validate location-related columns
        for loc_col_name, location_func in location_validation_rules.items():
            if loc_col_name in df.columns:
                try:
                    is_valid, color, highlight_col, message = location_func(row)
                    if is_valid:
                        print(f"Location validation successful: {loc_col_name} at row {idx + 2}")
                    else:
                        # Highlight the specific failing column
                        col_idx = df.columns.get_loc(highlight_col) + 1
                        ws.cell(row=idx + 2, column=col_idx).fill = highlight_fill_red if color == "red" else highlight_fill_yellow
                        print(f"Failed location validation: {loc_col_name} at row {idx + 2} - Reason: {message}")
                except Exception as e:
                    print(f"Error in location validation for {loc_col_name} at row {idx + 2}: {e}")

    # Save the workbook after validation and highlighting
    wb.save(output_file)
    print(f"Verification completed. Output saved as {output_file}")




# Argument parser to take file input and generate verified output
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Verify an Excel file against validation rules.")
    parser.add_argument("file_name", help="The name of the Excel file to verify")
    args = parser.parse_args()
    
    # Generate output file name with "Verified_" prefix
    input_file = args.file_name
    output_file = f"Verified_{os.path.basename(input_file)}"
    
    # Run verification
    verify_file(input_file, output_file)
